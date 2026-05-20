"""
步骤⑦：保质期填写 — 默认填 "3年"。

用法：
  python fill_shelf_life.py <excel_path> <column_letter> [--value "3年"] [--sheet <name>] [--dry-run]
  python fill_shelf_life.py "COA_output.xls" "K" --value "3年"
"""

import sys

try:
    import openpyxl
except ImportError:
    openpyxl = None


def fill_xlsx(path, col_letter, value, sheet_name=None, dry_run=False):
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    col_idx = openpyxl.utils.column_index_from_string(col_letter)
    count = 0
    for r in range(4, ws.max_row + 1):  # 从第4行开始，跳过表头
        cell = ws.cell(r, col_idx)
        if cell.value is None or str(cell.value).strip() == "":
            if not dry_run:
                cell.value = value
            count += 1

    if not dry_run:
        wb.save(path)
    wb.close()
    return count


def fill_xls_com(path, col_letter, value, sheet_name=None, dry_run=False):
    """.xls 文件用 COM 操作"""
    import win32com.client

    xl = win32com.client.Dispatch("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False

    wb = xl.Workbooks.Open(path)
    ws = wb.Sheets(sheet_name) if sheet_name else wb.Sheets(1)

    col_num = ord(col_letter.upper()) - ord("A") + 1
    last_row = ws.UsedRange.Rows.Count
    count = 0

    for r in range(4, last_row + 1):
        cell = ws.Cells(r, col_num)
        if cell.Value is None or str(cell.Value).strip() == "":
            if not dry_run:
                cell.Value = value
            count += 1

    if not dry_run:
        wb.Save()
    wb.Close()
    xl.Quit()
    return count


def main():
    if len(sys.argv) < 3:
        print("用法: python fill_shelf_life.py <excel_path> <column_letter> [--value '3年'] [--sheet <name>] [--dry-run]")
        sys.exit(1)

    path = sys.argv[1]
    col_letter = sys.argv[2].upper()
    value = "3年"
    sheet_name = None
    dry_run = "--dry-run" in sys.argv

    for i, arg in enumerate(sys.argv):
        if arg == "--value" and i + 1 < len(sys.argv):
            value = sys.argv[i + 1]
        if arg == "--sheet" and i + 1 < len(sys.argv):
            sheet_name = sys.argv[i + 1]

    is_xls = path.lower().endswith(".xls") and not path.lower().endswith(".xlsx")

    print(f"文件: {path}")
    print(f"列: {col_letter}, 填充值: '{value}'")
    if dry_run:
        print("[预览模式] 不保存")

    if is_xls:
        count = fill_xls_com(path, col_letter, value, sheet_name, dry_run)
    else:
        count = fill_xlsx(path, col_letter, value, sheet_name, dry_run)

    print(f"填充了 {count} 个空单元格 {'(未保存)' if dry_run else ''}")


if __name__ == "__main__":
    main()
