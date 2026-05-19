"""
步骤④：成分表含量合计检查。

检查每个 MSDS 文件每个成分表 sheet 的 D 列（Wt% 标准百分比）和 F 列（Actual Wt% 实际百分比）
合计是否约等于 100%。判定按优先级：

  1. D3/F3 有 =100-SUM(...) 公式 → OK（水补齐型）
  2. D 列全空、F 硬值合计≈100 → OK（源文件无 D 列）
  3. D 硬值合计≈100、F 列含公式 → OK（信任 Excel 计算）
  4. D 硬值合计≠100 → 异常（可能是混合物分量重复计）
  5. F 全硬值且合计≠100 → 异常

用法：
  python check_percent.py <base_dir> [--verbose]
  python check_percent.py "D:/MSDS/MSDS" --verbose
"""

import openpyxl
import os
import sys
import glob


def is_100_minus_sum(v):
    """检查是否 =100-SUM(...) 公式"""
    if not isinstance(v, str) or not v.startswith("="):
        return False
    return "100-SUM" in v.upper().replace(" ", "")


def check_sheet(ws):
    """检查单个成分表 sheet，返回 (ok, d_info, f_info)"""
    mr = ws.max_row
    d3 = ws.cell(3, 4).value
    f3 = ws.cell(3, 6).value
    d_formula_ok = is_100_minus_sum(d3)
    f_formula_ok = is_100_minus_sum(f3)

    if d_formula_ok and f_formula_ok:
        return True, "formula", "formula"

    d_sum = 0.0
    f_sum = 0.0
    d_hard = 0
    f_hard = 0
    f_formula_count = 0

    for r in range(3, mr + 1):
        dv = ws.cell(r, 4).value
        fv = ws.cell(r, 6).value
        if isinstance(dv, (int, float)):
            d_sum += dv
            d_hard += 1
        if isinstance(fv, (int, float)):
            f_sum += fv
            f_hard += 1
        elif isinstance(fv, str) and fv.startswith("="):
            f_formula_count += 1

    d_hard_ok = abs(d_sum - 100) < 0.001
    f_hard_ok = abs(f_sum - 100) < 0.001

    d_ok = d_formula_ok or d_hard_ok or d_hard == 0
    f_ok = f_formula_ok or f_hard_ok

    # F 含公式且 D 正确 → 信任 Excel 计算
    if not f_ok and f_formula_count > 0 and d_ok:
        f_ok = True

    # 生成诊断信息
    if d_formula_ok:
        d_info = "formula"
    elif d_hard == 0:
        d_info = "empty"
    else:
        d_info = f"sum={d_sum:.2f}"

    if f_formula_ok:
        f_info = "formula"
    elif f_formula_count > 0 and f_hard == 0:
        f_info = "all-formula"
    elif f_formula_count > 0:
        f_info = f"hard={f_sum:.2f}+{f_formula_count}fml"
    else:
        f_info = f"sum={f_sum:.2f}"

    return (d_ok and f_ok), d_info, f_info


def main():
    if len(sys.argv) < 2:
        print("用法: python check_percent.py <base_dir> [--verbose]")
        sys.exit(1)

    base = sys.argv[1]
    verbose = "--verbose" in sys.argv

    files = glob.glob(os.path.join(base, "**", "*_new.xlsx"), recursive=True)
    if not files:
        files = glob.glob(os.path.join(base, "**", "*.xlsx"), recursive=True)
    if not files:
        print(f"未找到 .xlsx 文件: {base}")
        sys.exit(1)

    print(f"检查 {len(files)} 个文件...\n")

    ok_count = 0
    issues = []

    for path in sorted(files):
        wb = openpyxl.load_workbook(path, data_only=False)
        fname = os.path.basename(path)

        for sname in wb.sheetnames[1:]:
            ok, d_info, f_info = check_sheet(wb[sname])
            if ok:
                ok_count += 1
                if verbose:
                    print(f"  OK  {fname}/{sname}")
            else:
                issues.append(f"  !! {fname}/{sname} | D={d_info}  F={f_info}")

        wb.close()

    print(f"\n正常: {ok_count} 个 sheet")
    if issues:
        print(f"异常: {len(issues)} 个 sheet")
        for i in issues:
            print(i)
    else:
        print("全部通过!")


if __name__ == "__main__":
    main()
