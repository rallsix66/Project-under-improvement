"""
步骤③：为 MSDS 文件 Sheet1 的产品名添加指向成分表 sheet 的内部超链接。

两种映射方式：
  A: Row 12 标签匹配 — Row 12 标签值与 sheet 名一一对应
  B: 位置映射 — Row 12 用 "第N页" 格式，sheet 名用 "第N款成分" 格式，按列序映射

用法：
  python hyperlink.py <folder_path> [--dry-run]
  python hyperlink.py "D:/MSDS/ICE-MSDS/已处理" --dry-run   # 预览不保存
  python hyperlink.py "D:/MSDS/ICE-MSDS/已处理"              # 执行
"""

import openpyxl
import os
import sys


LABEL_ROW = 12     # 页码标签行
TEXT_ROW = 13      # 产品名（超链接文字）行
START_COL = 3      # C 列 = 产品数据起始列


def add_hyperlinks(filepath, dry_run=False):
    wb = openpyxl.load_workbook(filepath)
    ws = wb[wb.sheetnames[0]]
    link_count = 0

    for col in range(START_COL, ws.max_column + 1):
        page_label = ws.cell(LABEL_ROW, col).value
        product_name = ws.cell(TEXT_ROW, col).value
        if not page_label or not product_name:
            continue

        target = None
        # 方式 A：标签值直接匹配 sheet 名
        if page_label in wb.sheetnames:
            target = page_label
        # 方式 B：位置映射（第N页 -> 第N个 sheet）
        elif col - START_COL + 1 < len(wb.sheetnames):
            target = wb.sheetnames[col - START_COL + 1]

        if target:
            cell = ws.cell(TEXT_ROW, col)
            cell.hyperlink = openpyxl.worksheet.hyperlink.Hyperlink(
                ref=cell.coordinate,
                location=f"'{target}'!A1",
                display=str(product_name),
            )
            link_count += 1

    if not dry_run:
        wb.save(filepath)
    wb.close()
    return link_count


def main():
    if len(sys.argv) < 2:
        print("用法: python hyperlink.py <folder_path> [--dry-run]")
        sys.exit(1)

    folder = sys.argv[1]
    dry_run = "--dry-run" in sys.argv

    files = sorted(
        f for f in os.listdir(folder) if f.endswith(".xlsx")
    )
    if not files:
        print(f"未找到 .xlsx 文件: {folder}")
        sys.exit(1)

    if dry_run:
        print(f"[预览模式] 找到 {len(files)} 个文件，不保存\n")
    else:
        print(f"找到 {len(files)} 个文件，开始处理...\n")

    # 先处理第一个文件
    first = files[0]
    path = os.path.join(folder, first)
    n = add_hyperlinks(path, dry_run=dry_run)
    print(f"  [{1}/{len(files)}] {first}: {n} links")

    if len(files) > 1:
        if not dry_run:
            resp = input("\n第一个文件已处理。继续处理其余文件？(y/n): ")
            if resp.lower() != "y":
                print("已取消")
                sys.exit(0)

        for i, fname in enumerate(files[1:], 2):
            path = os.path.join(folder, fname)
            n = add_hyperlinks(path, dry_run=dry_run)
            print(f"  [{i}/{len(files)}] {fname}: {n} links")

    print(f"\n{'[预览] ' if dry_run else ''}完成 {'(未保存)' if dry_run else ''}")


if __name__ == "__main__":
    main()
