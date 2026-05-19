# -*- coding: utf-8 -*-
"""MSDS生成脚本 v2：按品牌→生产单位分组，填写Sheet1基本资料+成分表
修正：sheet名不改、Row 10不填、Row 11单语言英文、地址加国家、生产单位英文翻译、
      Row 14-15单语言英文、多余列清理、成分表用replace_msds方法复制
"""
import os, re, json, shutil, time
import openpyxl
from copy import copy as pycopy
from collections import defaultdict
from deep_translator import GoogleTranslator

# ====== 路径配置（脚本所在 skill 目录） ======
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SKILL_DIR = os.path.dirname(SCRIPT_DIR)
TEMPLATE = os.path.join(SKILL_DIR, "templates", "MSDS模板.xlsx")
RESOURCES = os.path.join(SKILL_DIR, "resources")

# ====== 基本资料（从 resources/新建文本文档.txt 读取） ======
def load_delegate_info():
    info_file = os.path.join(RESOURCES, "新建文本文档.txt")
    info = {}
    if os.path.exists(info_file):
        with open(info_file, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if "：" in line:
                    k, v = line.split("：", 1)
                    info[k.strip()] = v.strip()
    return info

_DELEGATE = load_delegate_info()
DELEGATE_CN = _DELEGATE.get("委托单位", "")
DELEGATE_ADDR_CN = _DELEGATE.get("委托单位地址", "")
DELEGATE_EN = "Xiamen Bingliyan Cosmetics Co., Ltd."
DELEGATE_ADDR_EN = "Room 2503A, Block A, No. 2 Lianqian West Road, Siming District, Xiamen City, China"
CONTACT = _DELEGATE.get("联系人", "")
PHONE = _DELEGATE.get("电话", "")
EMAIL = _DELEGATE.get("邮箱", "")
BRAND_STD = {"ICE": "ICE LERSKIN", "CP": "CHIC.PEAK"}

# Row 11 语言设置（每次运行前确认）
ROW11_LANG = "英文"  # 单语言英文

# ====== 生产单位英文翻译（从 resources/mfr_en.json 加载） ======
def load_mfr_en():
    mfr_path = os.path.join(RESOURCES, "mfr_en.json")
    if os.path.exists(mfr_path):
        with open(mfr_path, "r", encoding="utf-8") as f:
            data = json.load(f)
            # 去掉 _description 说明字段
            return {k: v for k, v in data.items() if not k.startswith("_")}
    return {}

MFR_EN = load_mfr_en()
if not MFR_EN:
    MFR_EN = {"待确认": "TBD"}

CN_NUMS = ['一','二','三','四','五','六','七','八','九','十',
           '十一','十二','十三','十四','十五','十六','十七','十八','十九','二十']

def num_to_cn(n):
    """整数→中文数字（支持1-99）"""
    if n <= 20:
        return CN_NUMS[n - 1]
    tens = n // 10
    ones = n % 10
    if ones == 0:
        return f'{CN_NUMS[tens - 1]}十'
    return f'{CN_NUMS[tens - 1]}十{CN_NUMS[ones - 1]}'

PAGE_LABELS_CACHE = {}
def page_label(n):
    """第N页标签（n从1开始）"""
    if n not in PAGE_LABELS_CACHE:
        PAGE_LABELS_CACHE[n] = f'第{num_to_cn(n)}页'
    return PAGE_LABELS_CACHE[n]

TEMPLATE_PAGE_COUNT = 5  # 模板内置5个成分表sheet

# ====== 地址处理 ======
def ensure_country_cn(addr):
    """中文地址加中国前缀"""
    addr = addr.strip()
    if not addr:
        return addr
    if not addr.startswith('中国'):
        addr = '中国' + addr
    return addr

def ensure_country_en(addr):
    """英文地址末尾加 , China"""
    addr = addr.strip()
    if not addr:
        return addr
    if 'china' not in addr.lower():
        addr = addr + ', China'
    return addr

_ADDR_CACHE = {}

def translate_address_en(addr_cn):
    """中文地址 → 英文地址（Google翻译 + 后处理修正）"""
    addr = addr_cn.strip()
    if addr in _ADDR_CACHE:
        return _ADDR_CACHE[addr]

    src = addr if addr.startswith('中国') else '中国' + addr

    # 重试最多3次，间隔递增
    last_err = None
    for attempt in range(3):
        try:
            result = GoogleTranslator(source='zh-CN', target='en').translate(src)
            break
        except Exception as e:
            last_err = e
            if attempt < 2:
                time.sleep(2 * (attempt + 1))
    else:
        raise last_err

    # 后处理修正
    result = result.replace('self-reported', 'Self-declared')
    result = re.sub(r',\s*,+', ', ', result)
    result = re.sub(r'\s+', ' ', result).strip(', ')
    result = ensure_country_en(result)
    _ADDR_CACHE[addr] = result
    # 请求间短暂间隔，避免频率限制
    time.sleep(0.5)
    return result

# ====== 品名规范化 ======
def normalize_name_cn(cn, brand):
    std_brand = BRAND_STD[brand]
    cn = re.sub(r'(?i)\bchic\s*\.?\s*peak\b', 'CHIC.PEAK', cn)
    cn = re.sub(r'(?i)\bice\s+lerskin\b', 'ICE LERSKIN', cn)
    cn = re.sub(r'(?i)\bice\s?lerskin\b', 'ICE LERSKIN', cn)
    cn = re.sub(r'(ICE LERSKIN)([一-鿿])', r'\1 \2', cn)
    cn = re.sub(r'(CHIC\.PEAK)([一-鿿])', r'\1 \2', cn)
    # 去规格
    cn = re.sub(r'[（(]\s*\d+\.?\d*\s*(?:ml|g|kg|片)\s*(?:[+*x×]\s*\d+\.?\d*\s*(?:ml|g|kg|片))*[)）]\s*$', '', cn)
    cn = re.sub(r'\s*\d+\.?\d*\s*(?:ml|g|kg|片)\s*(?:[+*x×]\s*\d+\.?\d*\s*(?:ml|g|kg|片))*\s*$', '', cn)
    cn = re.sub(r'\s*\d+\.?\d*\s*(?:ml|g|kg|片)\s*(?:[+*x×]\s*\d+\.?\d*\s*(?:ml|g|kg|片))*\s*(?=[（(])', '', cn)
    cn = re.sub(r'\s*\d+\.?\d*\s*(?:ml|g|kg|片)\s*(?:[+*x×]\s*\d+\.?\d*\s*(?:ml|g|kg|片))*\s*-\s*', '-', cn)
    cn = re.sub(r'\s*-\s*$', '', cn)
    # 去版本描述
    cn = re.sub(r'\s*[（(][^）)]*(?:特供|升级|版)[^）)]*[）)]\s*$', '', cn)
    cn = re.sub(r'\s+', ' ', cn).strip()
    return cn

def normalize_name_en(en, brand):
    std_brand = BRAND_STD[brand]
    en = re.sub(r'(?i)\bchic\s*\.?\s*peak\b', 'CHIC.PEAK', en)
    en = re.sub(r'(?i)\bice\s+lerskin\b', 'ICE LERSKIN', en)
    en = re.sub(r'(?i)\bice\s?lerskin\b', 'ICE LERSKIN', en)
    if std_brand.upper() not in en.upper():
        en = f'{std_brand} {en}'
    en = re.sub(r'(ICE LERSKIN)\s+\1', r'\1', en)
    en = re.sub(r'(CHIC\.PEAK)\s+\1', r'\1', en)
    # 粘连拆分
    en = re.sub(r'([a-z])([A-Z])', r'\1 \2', en)
    en = re.sub(r'([A-Z]+)([A-Z][a-z])', r'\1 \2', en)
    en = re.sub(r'([A-Z])(ALL\b)', r'\1 \2', en)
    en = en.upper()
    en = re.sub(r'#(?=\s|$)', '', en)
    en = re.sub(r'\s+', ' ', en).strip()
    return en

# ====== 性状推断 ======
SHAPE_MAP = {
    '液体':'Liquid','透明液体':'Transparent Liquid','泡沫':'Foam','乳霜':'Cream',
    '慕斯':'Mousse','洁面':'Cleanser','凝胶':'Gel','膏状':'Paste','气雾':'Aerosol',
    '棒状':'Stick','手套型':'Glove Type','液状':'Liquid','乳液':'Emulsion',
    '精华':'Serum','喷雾':'Spray','粘稠液体':'Viscous Liquid','泥状':'Mud',
    '粉状':'Powder','待确认':'TBD',
}
COLOR_MAP = {
    '透明':'Transparent','白色':'White','白色+黑色':'White & Black','黑色':'Black',
    '半透明':'Translucent','淡黄色':'Light Yellow','淡紫色':'Light Purple',
    '淡粉色':'Light Pink','淡橘色':'Light Orange','蜜桃色':'Peach',
    '微白色':'Whitish','乳白色':'Milky White','无色透明':'Colorless Transparent',
    '透明泡沫':'Transparent Foam',
    '红色系':'Red Tone','橙色系':'Orange Tone','黄色系':'Yellow Tone',
    '蓝色系':'Blue Tone','紫色系':'Purple Tone','棕色系':'Brown Tone',
    '按色号':'By Shade','多色':'Multi-Color',
    '奶桃粉色':'Milky Peach Pink','草莓奶昔色':'Strawberry Milk Shake',
    '蜜瓜色':'Melon','树莓色':'Raspberry','烤奶棕色':'Toasted Brown',
    '象牙色':'Ivory','自然色':'Natural',
    '浆果色':'Berry','紫罗兰色':'Violet','冰茶色':'Iced Tea','日落色':'Sunset',
    '朦胧粉色':'Hazy Pink','象牙白':'Ivory White','中性色':'Neutral','暖肤色':'Warm Sand',
    '尘土玫瑰色':'Dusty Rose','丝绒蜜桃色':'Velvet Peach','裸杏色':'Nude Apricot',
    '肉桂拿铁色':'Cinnamon Latte','枣泥色':'Molasses Date','郁金香珊瑚色':'Tulip Coral',
    '待确认':'TBD',
}
ODOR_MAP = {
    '稍有气味':'Slight Odor','无':'Odorless',
    '特征性气味':'Characteristic Odor','待确认':'TBD',
}

# ====== CI 色号推断（与 COA 一致） ======
def infer_ci_color(ci_codes):
    if not ci_codes: return ''
    joined = ' '.join(ci_codes)
    if any(c in joined for c in ['CI15850','CI45410','CI73360','CI17200']): return '红色系'
    if any(c in joined for c in ['CI15985','CI45380']): return '橙色系'
    if any(c in joined for c in ['CI19140','CI47005']): return '黄色系'
    if any(c in joined for c in ['CI42090','CI77007']): return '蓝色系'
    if any(c in joined for c in ['CI60725','CI77742']): return '紫色系'
    if any(c in joined for c in ['CI77491','CI77492','CI77499']): return '棕色系'
    if 'CI77891' in joined and len(ci_codes) == 1: return '白色'
    if len(ci_codes) >= 2: return '多色'
    return ''

def read_ingredients_for_scent(ref_path, product_name_en, brand):
    """扫描成分文件的正确 sheet，返回 (has_fragrance, ci_codes)"""
    if not ref_path or not os.path.exists(ref_path):
        return False, []
    try:
        wb_ref = openpyxl.load_workbook(ref_path, data_only=True)
        name_up = product_name_en.upper().strip()
        ws_ref = None
        # A1 匹配
        for sn in wb_ref.sheetnames:
            ws = wb_ref[sn]
            a1 = str(ws.cell(row=1, column=1).value or '').strip().upper()
            if a1 and name_up in a1:
                ws_ref = ws
                break
        # 回退：品牌关键词
        if ws_ref is None:
            for sn in wb_ref.sheetnames:
                sn_up = sn.upper()
                if brand == 'ICE' and ('ICE' in sn_up or 'LITHE' in sn_up or 'DUAL' in sn_up):
                    ws_ref = wb_ref[sn]; break
                elif brand == 'CP' and 'CHIC' in sn_up:
                    ws_ref = wb_ref[sn]; break
        # 回退：第一个非 phantom
        if ws_ref is None:
            for sn in wb_ref.sheetnames:
                if 'LIP STICK#' not in sn.upper():
                    ws_ref = wb_ref[sn]; break
        if ws_ref is None:
            wb_ref.close()
            return False, []

        has_frag = False; ci_codes = []
        for r in range(3, ws_ref.max_row + 1):
            func = str(ws_ref.cell(r, 7).value or '').lower()
            if '香精' in func or 'fragrance' in func:
                has_frag = True
            for cc in [2, 3]:
                ci = str(ws_ref.cell(r, cc).value or '').strip()
                m = re.match(r'(CI\s*\d+)', ci, re.I)
                if m:
                    ci_codes.append(m.group(1).upper().replace(' ', ''))
                    break
        wb_ref.close()
        return has_frag, ci_codes
    except Exception:
        return False, []

# ====== 性状推断（AI/关键词，无外部依赖，与 COA 完全一致） ======
def get_shape_color_scent(brand, name_en, name_cn='', ref_path=''):
    """返回 (物理状态, 颜色, 气味) — Chinese。
    从成分表分析色素和香精推断，不依赖外部性状表。"""
    en = name_en.upper()

    has_frag, ci_codes = read_ingredients_for_scent(ref_path, name_en, brand)

    if brand == 'CP':
        odor = '稍有气味'

        # --- CP 物理状态 ---
        if 'LIP GLAZE' in en:
            state = '粘稠液体'
        elif 'LIP MUD' in en or 'BLUSH MUD' in en:
            state = '泥状'
        elif 'LIP JELLY' in en:
            state = '凝胶'
        elif 'LIQUID BLUSH' in en:
            state = '液体'
        elif 'LIP STICK' in en and 'LIP LINER' in en:
            state = '棒状'
        elif 'LIP STICK' in en:
            state = '棒状'
        elif 'CUSHION' in en:
            state = '液状'
        elif 'SETTING COMPACT' in en:
            state = '粉状'
        else:
            state = '膏状'

        # --- CP 颜色 ---
        color = ''
        # 从中文品名提取颜色
        color_from_cn = re.search(r'(#[0-9]{2}|[0-9]{2}#)[（(]?\s*([一-鿿]+[色白])', name_cn)
        if color_from_cn:
            color = color_from_cn.group(2)
        # LIP STICK 英文色号翻译
        if not color and 'LIP STICK' in en:
            lip_colors_en = {
                'DUSTY ROSE': '尘土玫瑰色', 'VELVET PEACH': '丝绒蜜桃色',
                'NUDE APRICOT': '裸杏色', 'CINNAMON LATTE': '肉桂拿铁色',
                'MOLASSES DATE': '枣泥色', 'TULIP CORAL': '郁金香珊瑚色',
            }
            for ek, cn_color in lip_colors_en.items():
                if ek in en:
                    color = cn_color; break
        # CI 色号推断
        if not color:
            color = infer_ci_color(ci_codes)
        if not color:
            color = '多色'

    else:
        # ICE 品牌
        state = ''
        if any(w in en for w in ['TONER','LOTION','WATER']): state = '液体'
        if 'SERUM' in en: state = '精华'
        if 'SPRAY' in en and 'STICK' not in en: state = '气雾' if 'SUN' in en else '液体'
        if 'STICK' in en: state = '棒状'
        if 'CREAM' in en and 'EYE' in en: state = '乳霜'
        if 'MOUSSE' in en or 'CLEANSER' in en or 'CLEANSING' in en: state = '慕斯'
        if 'GEL' in en: state = '凝胶'
        if 'MASK' in en or 'MUD' in en: state = '膏状'
        if 'BATH GLOVES' in en: state = '手套型'
        if 'OIL' in en or 'ELIXIR' in en: state = '液体'
        if 'HYDRO-GLOW LIP' in en or 'LIP' in en:
            if 'OIL' in en: state = '液体'
            elif 'BALM' in en: state = '膏状'
        if not state: state = '待确认'

        color = ''
        if any(w in en for w in ['CLEANSER','MOUSSE','CLEANSING']): color = '透明'
        elif state in ('液体','精华'): color = '透明'
        elif state == '棒状': color = '白色'
        if not color: color = ''

        odor = '稍有气味' if has_frag else ''

        # ICE 历史数据覆盖
        if 'GLYCOLIC' in en: state, color, odor = '液体', '透明', '稍有气味'
        if 'AIR LIGHT' in en and 'SUN' in en: state, color, odor = '棒状', '白色', '稍有气味'
        if 'ACNE' in en and 'MOUSSE' in en: state, color, odor = '泡沫', '透明', '稍有气味'
        if 'ANTI-GRAVITY' in en: state, color, odor = '乳霜', '白色', '稍有气味'
        if 'DUAL-COLOR' in en: state, color, odor = '膏状', '白色+黑色', '稍有气味'
        if 'GLACIER' in en: state, color, odor = '洁面', '透明', '稍有气味'
        if 'SUNSPRITZ' in en: state, color, odor = '喷雾', '透明', '稍有气味'
        if 'MULTI-EFFECT WHITENING' in en: state, color, odor = '气雾', '乳白色', '稍有气味'
        if 'SEA FENNEL' in en: state, color, odor = '液体', '透明', '稍有气味'
        if 'UV PLUS' in en: state, color, odor = '乳液', '白色', '稍有气味'
        if 'BHA' in en and 'PORE' in en: state, color, odor = '凝胶', '半透明', '无'
        if 'C E FERULIC' in en: state, color, odor = '精华', '透明', '无'
        if 'HYDRO-GLOW LIP' in en:
            state = '膏状'
            if 'BANANA' in en: color = '淡黄色'
            elif 'GRAPE' in en and 'GRAPEFRUIT' not in en: color = '淡紫色'
            elif 'GRAPEFRUIT' in en: color = '淡粉色'
            elif 'PEACH' in en: color = '蜜桃色'
            odor = '稍有气味'
        if 'DAY & NIGHT' in en: state, color, odor = '液体', '淡橘色', '稍有气味'
        if 'EVEN-TONE' in en: state, color, odor = '精华', '透明', '稍有气味'
        if 'BATH GLOVES' in en: state, color, odor = '手套型', '白色', '稍有气味'
        if 'TRIPLE ACID' in en: state, color, odor = '喷雾', '透明', '无'
        if 'PRECISION ACNE' in en: state, color, odor = '液体', '微白色', '稍有气味'
        if 'LITHE' in en: state, color, odor = '液体', '透明', '特征性气味'

        if not state: state = '待确认'
        if not odor: odor = '待确认'

    if not color: color = '待确认'
    if not state: state = '待确认'
    if odor in ('轻微气味', '有香精'):
        odor = '稍有气味'
    return state, color, odor

def translate_shape_color_en(state_cn, color_cn, odor_cn):
    """性状→英文（Row 14-15 单语言）"""
    state_en = SHAPE_MAP.get(state_cn, state_cn)
    color_en = COLOR_MAP.get(color_cn, color_cn)
    odor_en = ODOR_MAP.get(odor_cn, odor_cn)
    return f"{color_en} {state_en}".strip(), odor_en

# ====== 成分表复制（replace_msds 方法） ======
def copy_cell_style(src_cell, dst_cell):
    """复制单元格样式"""
    if src_cell.has_style:
        dst_cell.font = pycopy(src_cell.font)
        dst_cell.fill = pycopy(src_cell.fill)
        dst_cell.border = pycopy(src_cell.border)
        dst_cell.alignment = pycopy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format

def copy_ingredients(ws_msds, ref_path, product_name_en, brand):
    """从参考.xlsm文件复制成分数据到MSDS sheet。
    使用 skill replace_msds_batch.py 的 A1 匹配方法找正确的 sheet，
    然后 unmerge→清空→复制数据+样式+合并→清多余行。
    """
    wb_ref = openpyxl.load_workbook(ref_path, data_only=False)
    name_up = product_name_en.upper().strip()

    # 1. A1 品名匹配（与 replace_msds_batch.py 的 find_reference_sheet 一致）
    ws_ref = None
    for sn in wb_ref.sheetnames:
        ws = wb_ref[sn]
        a1 = str(ws.cell(row=1, column=1).value or '').strip().upper()
        if a1 and name_up in a1:
            ws_ref = ws
            break

    # 2. 回退：品牌关键词匹配 sheet 名
    if ws_ref is None:
        for sn in wb_ref.sheetnames:
            sn_up = sn.upper()
            if brand == 'ICE' and ('ICE' in sn_up or 'LITHE' in sn_up or 'DUAL' in sn_up):
                ws_ref = wb_ref[sn]
                break
            elif brand == 'CP' and 'CHIC' in sn_up:
                ws_ref = wb_ref[sn]
                break

    # 3. 回退：第一个非 phantom sheet（跳过 LIP STICK# 模板残留）
    if ws_ref is None:
        for sn in wb_ref.sheetnames:
            if 'LIP STICK#' not in sn.upper():
                ws_ref = wb_ref[sn]
                break

    if ws_ref is None:
        wb_ref.close()
        return False

    # 4. 解除 MSDS 合并
    for mc in list(ws_msds.merged_cells.ranges):
        ws_msds.unmerge_cells(str(mc))

    # 5. 读取参考表合并
    ref_merges = list(ws_ref.merged_cells.ranges)

    # 6. 清空区域
    need_rows = max(ws_msds.max_row, ws_ref.max_row)
    need_cols = max(ws_msds.max_column, ws_ref.max_column)
    for r in range(1, need_rows + 1):
        for c in range(1, need_cols + 1):
            ws_msds.cell(row=r, column=c).value = None

    # 7. 复制参考表数据+样式+行高
    for r in range(1, ws_ref.max_row + 1):
        if ws_ref.row_dimensions[r].height:
            ws_msds.row_dimensions[r].height = ws_ref.row_dimensions[r].height
        for c in range(1, ws_ref.max_column + 1):
            src = ws_ref.cell(row=r, column=c)
            dst = ws_msds.cell(row=r, column=c)
            dst.value = src.value
            copy_cell_style(src, dst)

    # 8. 复制列宽
    for c in range(1, ws_ref.max_column + 1):
        letter = openpyxl.utils.get_column_letter(c)
        if ws_ref.column_dimensions[letter].width:
            ws_msds.column_dimensions[letter].width = ws_ref.column_dimensions[letter].width

    # 9. 添加参考表合并
    for mc in ref_merges:
        ws_msds.merge_cells(str(mc))

    # 10. 清除多余行（值+样式清空）
    empty_fill = openpyxl.styles.PatternFill(fill_type=None)
    empty_font = openpyxl.styles.Font()
    empty_border = openpyxl.styles.Border()
    empty_align = openpyxl.styles.Alignment()
    for r in range(ws_ref.max_row + 1, need_rows + 1):
        for c in range(1, need_cols + 1):
            cell = ws_msds.cell(row=r, column=c)
            cell.value = None
            cell.fill = pycopy(empty_fill)
            cell.font = pycopy(empty_font)
            cell.border = pycopy(empty_border)
            cell.alignment = pycopy(empty_align)

    wb_ref.close()
    return True

# ====== Sheet1 表格延伸（产品>5时） ======
def extend_sheet1(wb, ws, n_products):
    """产品超过模板5个时，延伸Sheet1合并单元格和边框样式"""
    if n_products <= TEMPLATE_PAGE_COUNT:
        return

    last_col_idx = 2 + n_products
    last_col_letter = openpyxl.utils.get_column_letter(last_col_idx)
    tmpl_data_col = 2 + TEMPLATE_PAGE_COUNT  # 模板最后一个数据列 (G=7)
    merge_boundary_col = 11  # 模板合并边界列 K

    # 产品 ≤9（模板K列范围）：只复制样式到新列，不缩合并单元格
    if last_col_idx <= merge_boundary_col:
        for r in range(1, ws.max_row + 1):
            src_cell = ws.cell(row=r, column=tmpl_data_col)
            for col_idx in range(tmpl_data_col + 1, last_col_idx + 1):
                dst_cell = ws.cell(row=r, column=col_idx)
                copy_cell_style(src_cell, dst_cell)
        tmpl_letter = openpyxl.utils.get_column_letter(tmpl_data_col)
        tmpl_width = ws.column_dimensions[tmpl_letter].width
        for col_idx in range(tmpl_data_col + 1, last_col_idx + 1):
            letter = openpyxl.utils.get_column_letter(col_idx)
            if tmpl_width:
                ws.column_dimensions[letter].width = tmpl_width
        return

    # 保存 K 列右边框（合并区域右边界线），延伸后移到新末列
    k_right_borders = {}
    for r in range(1, ws.max_row + 1):
        k_right_borders[r] = pycopy(ws.cell(r, merge_boundary_col).border.right)

    # 1. 延伸合并单元格：Rows 1-11
    for mc in list(ws.merged_cells.ranges):
        mc_str = str(mc)
        if mc_str.startswith('B1:'):
            ws.unmerge_cells(mc_str)
            ws.merge_cells(f'B1:{last_col_letter}1')
        elif mc_str.startswith('C') and ':K' in mc_str:
            parts = mc_str.split(':')
            end_cell = parts[1]
            r_match = re.search(r'\d+', end_cell)
            if r_match:
                r = int(r_match.group())
                if 2 <= r <= 11:
                    ws.unmerge_cells(mc_str)
                    ws.merge_cells(f'C{r}:{last_col_letter}{r}')

    # 2. 在新列复制样式（从模板数据末列 G 复制到新列）
    #    覆盖所有行（含 Row 17-19）
    for r in range(1, ws.max_row + 1):
        src_cell = ws.cell(row=r, column=tmpl_data_col)
        for col_idx in range(tmpl_data_col + 1, last_col_idx + 1):
            dst_cell = ws.cell(row=r, column=col_idx)
            copy_cell_style(src_cell, dst_cell)

    # 3. 将 K 列的右边框移到新的末列
    for r in range(1, ws.max_row + 1):
        saved = k_right_borders.get(r)
        # 新末列：套上原来的粗右边框
        if saved and saved.style:
            new_cell = ws.cell(row=r, column=last_col_idx)
            b = new_cell.border
            new_cell.border = openpyxl.styles.Border(
                left=b.left, right=saved, top=b.top, bottom=b.bottom,
                diagonal=b.diagonal, outline=b.outline,
                vertical=b.vertical, horizontal=b.horizontal)
        # 旧 K 列：去掉右边框（已变为内部列）
        old_cell = ws.cell(row=r, column=merge_boundary_col)
        b = old_cell.border
        old_cell.border = openpyxl.styles.Border(
            left=b.left, right=openpyxl.styles.Side(style=None),
            top=b.top, bottom=b.bottom,
            diagonal=b.diagonal, outline=b.outline,
            vertical=b.vertical, horizontal=b.horizontal)

    # 4. 设置新列列宽（复制 G 列宽）
    tmpl_letter = openpyxl.utils.get_column_letter(tmpl_data_col)
    tmpl_width = ws.column_dimensions[tmpl_letter].width
    for col_idx in range(tmpl_data_col + 1, last_col_idx + 1):
        letter = openpyxl.utils.get_column_letter(col_idx)
        if tmpl_width:
            ws.column_dimensions[letter].width = tmpl_width

# ====== 公式自动修正 ======
def auto_fix_formulas(wb):
    """遍历成分表 sheet，检查并修正 =100-SUM 公式范围，反馈需人工的异常。
    返回 (fixed: list, manual: list)"""
    fixed = []
    manual = []

    for sname in wb.sheetnames[1:]:
        ws = wb[sname]
        mr = ws.max_row
        if mr < 3:
            continue

        # 找 D 列 / F 列实际有数据的最后一行
        def last_data_row(col):
            for r in range(mr, 2, -1):
                v = ws.cell(r, col).value
                if v is not None and str(v).strip() != '':
                    return r
            return 0

        d_last = last_data_row(4)
        f_last = last_data_row(6)

        d3v = ws.cell(3, 4).value
        f3v = ws.cell(3, 6).value
        d3_str = str(d3v or '')
        f3_str = str(f3v or '')

        d_done = False
        f_done = False

        # --- 修正 D3 =100-SUM 公式（范围到实际数据末行） ---
        d_range_ok = False
        if d3_str.startswith('=') and '100-SUM' in d3_str.upper():
            m = re.search(r'SUM\((\w+):(\w+)\)', d3_str, re.I)
            if m:
                end_row = int(re.search(r'\d+', m.group(2)).group())
                if end_row == d_last:
                    d_range_ok = True
                elif d_last > 0:
                    new_formula = f'=100-SUM({m.group(1)}:D{d_last})'
                    ws.cell(3, 4).value = new_formula
                    d_done = True

        # --- 修正 F3 =100-SUM 公式 ---
        f_range_ok = False
        if f3_str.startswith('=') and '100-SUM' in f3_str.upper():
            m = re.search(r'SUM\((\w+):(\w+)\)', f3_str, re.I)
            if m:
                end_row = int(re.search(r'\d+', m.group(2)).group())
                if end_row == f_last:
                    f_range_ok = True
                elif f_last > 0:
                    new_formula = f'=100-SUM({m.group(1)}:F{f_last})'
                    ws.cell(3, 6).value = new_formula
                    f_done = True

        # --- 如果 D3 没有 =100-SUM 公式，检查硬值合计 ---
        if not d_done and not d_range_ok and not d3_str.startswith('='):
            d_sum = sum(
                ws.cell(r, 4).value for r in range(3, mr + 1)
                if isinstance(ws.cell(r, 4).value, (int, float))
            )
            if abs(d_sum - 100) >= 0.001 and d_sum > 0:
                d_has_fml = any(
                    isinstance(ws.cell(r, 4).value, str) and str(ws.cell(r, 4).value).startswith('=')
                    for r in range(3, mr + 1)
                )
                if d_has_fml:
                    manual.append(f"{sname}: D列含公式但合计={d_sum:.2f}，需人工")
                else:
                    manual.append(f"{sname}: D列全硬值合计={d_sum:.2f}，无公式可调整，请人工确认")

        # --- 如果 F3 没有 =100-SUM 公式，检查硬值合计 ---
        if not f_done and not f_range_ok and not f3_str.startswith('='):
            f_sum = sum(
                ws.cell(r, 6).value for r in range(3, mr + 1)
                if isinstance(ws.cell(r, 6).value, (int, float))
            )
            # F 列含公式则信任 Excel（已在 check_totals 处理）
            f_has_fml = any(
                isinstance(ws.cell(r, 6).value, str) and str(ws.cell(r, 6).value).startswith('=')
                for r in range(3, mr + 1)
            )
            if f_has_fml:
                pass  # 信任公式
            elif abs(f_sum - 100) >= 0.001 and f_sum > 0:
                manual.append(f"{sname}: F列全硬值合计={f_sum:.2f}，无公式可调整，请人工确认")

        if d_done or f_done:
            parts = []
            if d_done: parts.append('D已修正')
            if f_done: parts.append('F已修正')
            fixed.append(f"{sname}: {', '.join(parts)}")

    return fixed, manual


# ====== 主生成函数 ======
def create_msds(group_key, products, base_dir, output_dir):
    brand, mfr = group_key.split('|')
    brand_std = BRAND_STD[brand]
    n = len(products)

    mfr_en = MFR_EN.get(mfr, mfr)
    addr_cn = products[0]['address']
    addr_en = translate_address_en(addr_cn)
    addr_cn_full = ensure_country_cn(addr_cn)

    out_name = f"MSDS_{brand}_{mfr}.xlsx"
    out_path = os.path.join(output_dir, brand, out_name)
    os.makedirs(os.path.dirname(out_path), exist_ok=True)

    ref_dir = os.path.join(base_dir, f"{brand}成分")

    shutil.copyfile(TEMPLATE, out_path)
    wb = openpyxl.load_workbook(out_path)

    # ===== 删除多余成分表sheet（产品<5时） =====
    for i in range(n, TEMPLATE_PAGE_COUNT):
        sheet_name = f'第{num_to_cn(i + 1)}款成分'
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    # ===== 产品超过5个时创建新sheet =====
    for i in range(TEMPLATE_PAGE_COUNT, n):
        new_name = f'第{num_to_cn(i + 1)}款成分'
        if new_name not in wb.sheetnames:
            last_template = f'第{num_to_cn(TEMPLATE_PAGE_COUNT)}款成分'
            if last_template in wb.sheetnames:
                ws_src = wb[last_template]
                wb.copy_worksheet(ws_src).title = new_name

    # ===== 产品超过5个时延伸Sheet1表格 =====
    extend_sheet1(wb, wb[wb.sheetnames[0]], n)

    # ===== Sheet1 填写 =====
    ws = wb[wb.sheetnames[0]]

    # Rows 2-7, 9: 基本资料（模板标注"中文/英文"的行用 / 分隔）
    ws.cell(2, 3).value = f"{DELEGATE_CN} / {DELEGATE_EN}"
    ws.cell(3, 3).value = f"{DELEGATE_ADDR_CN} / {DELEGATE_ADDR_EN}"
    ws.cell(4, 3).value = f"{mfr} / {mfr_en}"
    ws.cell(5, 3).value = f"{addr_cn_full} / {addr_en}"
    ws.cell(6, 3).value = CONTACT
    ws.cell(7, 3).value = PHONE
    # Row 8 不填（紧急联系话语）
    ws.cell(9, 3).value = EMAIL
    # Row 10 不填（品牌准确名称）
    # Row 11 填语言
    ws.cell(11, 3).value = ROW11_LANG

    # Rows 12-15: 产品信息
    for i, p in enumerate(products):
        col = 3 + i
        cn = p['name_cn_norm']
        en = p['name_en_norm']

        ws.cell(12, col).value = page_label(i + 1)
        # Row 13: 模板标注"中文/英文" → 双语
        ws.cell(13, col).value = f"{cn} / {en}"

        fname = p.get('filename', '')
        ref_path = os.path.join(ref_dir, fname) if fname else ''

        state_cn, color_cn, odor_cn = get_shape_color_scent(brand, en, cn, ref_path)
        appear_en, odor_en = translate_shape_color_en(state_cn, color_cn, odor_cn)
        # Row 14-15: 模板未标注"中文/英文" → 单语言英文
        ws.cell(14, col).value = appear_en
        ws.cell(15, col).value = odor_en

    # 清除多余列（产品 < 模板列数时）
    template_data_cols = 2 + TEMPLATE_PAGE_COUNT  # 模板数据结束列
    total_data_cols = 2 + n  # 实际数据结束列
    for i in range(n, max(TEMPLATE_PAGE_COUNT, n) + 1):
        col = 3 + i
        # 清除 Row 12-15
        for r in [12, 13, 14, 15]:
            ws.cell(r, col).value = None
    # 如果产品 < 5，清除模板多余列在合并区域内的残留
    if n < TEMPLATE_PAGE_COUNT:
        for r in [12, 13, 14, 15]:
            for c in range(3 + n, 3 + TEMPLATE_PAGE_COUNT):
                ws.cell(r, c).value = None
        # Row 12 模板有6列(C-H)，清到H列
        for c in range(3 + n, 9):  # 9 = H列+1
            for r in [12, 13, 14, 15]:
                ws.cell(r, c).value = None

    # Row 12 多余列也清（模板可能有多于5的标签列）
    for col in range(3 + n, ws.max_column + 1):
        page_val = ws.cell(12, col).value
        if page_val and isinstance(page_val, str) and page_val.startswith('第'):
            ws.cell(12, col).value = None

    # ===== 成分表复制 =====
    for i, p in enumerate(products):
        sheet_name = f'第{num_to_cn(i + 1)}款成分'
        fname = p.get('filename', '')
        ref_path = os.path.join(ref_dir, fname) if fname else ''

        if ref_path and os.path.exists(ref_path):
            ok = copy_ingredients(wb[sheet_name], ref_path, p['name_en_norm'], brand)
            if not ok:
                print(f"  Warning: copy failed for {fname}")
        else:
            print(f"  Warning: ref not found: {fname}")

    # ===== Sheet1 超链接：Row 13 产品名 → 对应成分表 sheet =====
    ing_sheets = wb.sheetnames[1:]  # 跳过 Sheet1
    for i in range(n):
        col = 3 + i
        sheet_name = f'第{num_to_cn(i + 1)}款成分'
        if sheet_name in ing_sheets:
            cell = ws.cell(13, col)
            cell.hyperlink = openpyxl.worksheet.hyperlink.Hyperlink(
                ref=cell.coordinate,
                location=f"'{sheet_name}'!A1",
                display=str(cell.value)
            )

    # ===== 公式自动修正 =====
    fixed, manual = auto_fix_formulas(wb)
    if fixed or manual:
        print(f"  公式修正: {len(fixed)} fixed, {len(manual)} manual")

    wb.save(out_path)
    wb.close()
    return out_path


# ====== 含量合计检查 ======
def check_totals(file_path):
    """检查 MSDS 文件各成分表 D/F 列合计是否 ≈ 100%

    判定规则（按优先级）：
    1. D3/F3 有 =100-SUM(...) 公式 → OK（水补齐型，自动保证100%）
    2. D 全空，F 硬值合计 ≈100 → OK（源文件无 D 列）
    3. D 硬值合计 ≈100，F 含公式 → OK（公式在 Excel 中自动计算正确）
    4. 其他情况：
       - D 硬值合计 ≠100 → 异常（可能是混合物分量重复计数）
       - F 全硬值且合计 ≠100 → 异常
    """
    wb = openpyxl.load_workbook(file_path, data_only=False)
    fname = os.path.basename(file_path)

    ok_count = 0
    issues = []

    def is_100_minus_sum(v):
        if not isinstance(v, str) or not v.startswith('='):
            return False
        return '100-SUM' in v.upper().replace(' ', '')

    for sname in wb.sheetnames[1:]:
        ws = wb[sname]
        mr = ws.max_row

        d3 = ws.cell(3, 4).value
        f3 = ws.cell(3, 6).value

        d_formula_ok = is_100_minus_sum(d3)
        f_formula_ok = is_100_minus_sum(f3)

        if d_formula_ok and f_formula_ok:
            ok_count += 1
            continue

        d_sum = 0.0
        f_sum = 0.0
        d_hard = 0
        f_hard = 0
        f_formula_count = 0
        for r in range(3, mr + 1):
            dv = ws.cell(r, 4).value
            fv = ws.cell(r, 6).value
            if isinstance(dv, (int, float)):
                d_sum += dv
                d_hard += 1
            if isinstance(fv, (int, float)):
                f_sum += fv
                f_hard += 1
            elif isinstance(fv, str) and fv.startswith('='):
                f_formula_count += 1

        d_hard_ok = abs(d_sum - 100) < 0.001
        f_hard_ok = abs(f_sum - 100) < 0.001

        # D OK: formula-OK, or hard-sum=100, or no D values at all
        d_ok = d_formula_ok or d_hard_ok or d_hard == 0
        # F OK: formula-OK, or hard-sum=100, or has formulas AND D is OK
        f_ok = f_formula_ok or f_hard_ok
        if not f_ok and f_formula_count > 0 and d_ok:
            f_ok = True  # F formulas auto-calculate correctly when D is correct

        if d_ok and f_ok:
            ok_count += 1
        else:
            # Build detailed info
            if d_formula_ok:
                d_info = 'formula'
            elif d_hard == 0:
                d_info = 'empty'
            else:
                d_info = f'sum={d_sum:.2f}'
            if f_formula_ok:
                f_info = 'formula'
            elif f_formula_count > 0 and f_hard == 0:
                f_info = 'all-formula'
            elif f_formula_count > 0:
                f_info = f'hard={f_sum:.2f}+{f_formula_count}fml'
            else:
                f_info = f'sum={f_sum:.2f}'
            issues.append(
                f"  {fname}/{sname} | D={d_info}  F={f_info}"
            )

    wb.close()
    return ok_count, issues

# ====== Main ======
if __name__ == '__main__':
    import sys
    if len(sys.argv) < 3:
        print("用法: python msds_fill.py <数据目录> <输出目录> [语言]")
        print("  数据目录: 含 CP成分/、ICE成分/、厦门外贸商品对应生产方.xlsx")
        print("  输出目录: MSDS 输出位置")
        print("  语言: Row 11 语言设置，默认'英文'")
        sys.exit(1)

    BASE = sys.argv[1]
    OUTPUT_DIR = sys.argv[2]
    if len(sys.argv) >= 4:
        ROW11_LANG = sys.argv[3]

    # 加载生产方对照表（与 COA 共用 match_mfr）
    mfr_table = os.path.join(BASE, "厦门外贸商品对应生产方.xlsx")
    if not os.path.exists(mfr_table):
        print(f"错误: 找不到生产方对照表 {mfr_table}")
        sys.exit(1)

    from coa_fill import match_mfr

    wb_mfr = openpyxl.load_workbook(mfr_table, data_only=True)
    ws_mfr = wb_mfr.active
    MFR = []
    for r in range(2, ws_mfr.max_row + 1):
        code = str(ws_mfr.cell(r, 1).value or '').strip()
        name = str(ws_mfr.cell(r, 2).value or '').strip()
        mfr  = str(ws_mfr.cell(r, 3).value or '').strip()
        addr = str(ws_mfr.cell(r, 4).value or '').strip()
        if name:
            MFR.append((code, name, mfr, addr))
    wb_mfr.close()
    print(f"生产方表: {len(MFR)}条")

    groups = defaultdict(list)

    for brand in ['CP', 'ICE']:
        brand_dir = os.path.join(BASE, f"{brand}成分")
        if not os.path.exists(brand_dir):
            continue
        for fname in os.listdir(brand_dir):
            if fname.startswith('~$') or not fname.endswith('.xlsm'):
                continue
            fpath = os.path.join(brand_dir, fname)
            try:
                wb = openpyxl.load_workbook(fpath, data_only=True)
                ws = wb[wb.sheetnames[-1]]
                en_name = str(ws.cell(1, 1).value or '').strip()
                wb.close()
                if not en_name:
                    continue

                # 用 COA 同款匹配逻辑
                result, score = match_mfr(en_name, brand, MFR)
                if result and score >= 3:
                    code, cn_name, mfr, addr = result
                else:
                    cn_name = en_name
                    mfr = '待确认'
                    addr = ''

                if mfr:
                    cn = normalize_name_cn(cn_name, brand)
                    en = normalize_name_en(en_name, brand)
                    key = f"{brand}|{mfr}"
                    groups[key].append({
                        'brand': brand, 'name_cn': cn, 'name_en': en,
                        'name_cn_norm': cn, 'name_en_norm': en,
                        'filename': fname, 'address': addr,
                        'manufacturer': mfr
                    })
            except Exception as e:
                print(f"  跳过 {fname}: {e}")

    print(f"分组结果: {len(groups)} 组")
    for key in sorted(groups.keys()):
        print(f"  {key}: {len(groups[key])} 个产品")

    # 预翻译所有唯一地址
    unique_addrs = sorted({p['address'].strip() for prods in groups.values() for p in prods if p.get('address')})
    print(f"\n{'='*60}")
    print(f"预翻译 {len(unique_addrs)} 个唯一地址...")
    print(f"{'='*60}")
    for addr in unique_addrs:
        try:
            en = translate_address_en(addr)
            print(f"  OK: {addr[:30]}... → {en[:60]}...")
        except Exception as e:
            print(f"  FAIL: {addr[:30]}... → {e}")
    print()

    total_ok = 0
    total_issues = []
    for key in sorted(groups.keys()):
        prods = groups[key]
        brand, mfr = key.split('|')
        print(f"\n{'='*60}")
        print(f"[{brand}] {mfr}: {len(prods)} products")
        print(f"{'='*60}")
        for p in prods:
            print(f"  {p['name_en_norm'][:70]}")
            ref_path = os.path.join(BASE, f"{brand}成分", p.get('filename', '')) if p.get('filename') else ''
            s, c, o = get_shape_color_scent(p['brand'], p['name_en_norm'], p.get('name_cn_norm', ''), ref_path)
            ae, oe = translate_shape_color_en(s, c, o)
            print(f"    State={s}  Color={c}  Odor={o} → {ae} / {oe}")

        out = create_msds(key, prods, BASE, OUTPUT_DIR)
        print(f"  Output: {os.path.basename(out)}")

        ok, issues = check_totals(out)
        total_ok += ok
        total_issues.extend(issues)
        print(f"  含量: {ok}/{len(prods)} OK")
        if issues:
            for i in issues:
                print(f"    ISSUE: {i}")

    print(f"\n{'='*60}")
    print(f"全部完成: {total_ok} sheets OK")
    if total_issues:
        print(f"异常: {len(total_issues)} 个sheet")
        for i in total_issues:
            print(f"  {i}")
    else:
        print("含量全部通过!")
