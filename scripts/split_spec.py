"""
步骤⑤：产品名拆分 — 从中文品名中分离产品名和规格。

规则：
  - 规格关键词：ml, g, 片, 支, 瓶, 盒, 包, 条, kg, L, oz, 对, 套, 粒
  - 规格前有空格 → 在空格处拆分
  - 无空格 → 在规格关键词处拆分
  - 产品名不含规格信息

用法：
  python split_spec.py <excel_path> <name_column> [--sheet <name>] [--dry-run]
"""

import openpyxl
import re
import sys


SPEC_PATTERN = re.compile(
    r'(\d+\.?\d*\s*(?:ml|g|片|支|瓶|盒|包|条|kg|L|oz|对|套|粒|mm|cm|inch|oz))',
    re.IGNORECASE,
)


def split_name(full_name):
    """拆分品名为 (产品名, 规格)"""
    if not full_name:
        return "", ""

    full_name = str(full_name).strip()

    # 规格在括号中 → 直接提取
    bracket_match = re.search(r'[（(]([^）)]+)[）)]$', full_name)
    if bracket_match:
        spec = bracket_match.group(1)
        name = full_name[:bracket_match.start()].strip()
        return name, spec

    # 正则匹配规格关键词
    match = SPEC_PATTERN.search(full_name)
    if not match:
        return full_name, ""

    spec_start = match.start()
    # 如果规格前有空格，从空格处拆分
    space_before = full_name.rfind(" ", 0, spec_start)
    if space_before > 0:
        name = full_name[:space_before].strip()
        spec = full_name[space_before + 1:].strip()
    else:
        name = full_name[:spec_start].strip()
        spec = full_name[spec_start:].strip()

    return name, spec


def main():
    if len(sys.argv) < 3:
        print("用法: python split_spec.py <excel_path> <name_column_letter> [--sheet <name>] [--dry-run]")
        sys.exit(1)

    path = sys.argv[1]
    name_col = sys.argv[2].upper()
    sheet_name = None
    dry_run = "--dry-run" in sys.argv

    for i, arg in enumerate(sys.argv):
        if arg == "--sheet" and i + 1 < len(sys.argv):
            sheet_name = sys.argv[i + 1]

    wb = openpyxl.load_workbook(path)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    col_idx = openpyxl.utils.column_index_from_string(name_col)
    mr = ws.max_row

    print(f"文件: {path}")
    print(f"Sheet: {ws.title}, 列: {name_col}, 行数: {mr}")
    if dry_run:
        print("[预览模式] 不保存\n")

    for r in range(1, mr + 1):
        val = ws.cell(r, col_idx).value
        if not val:
            continue
        name, spec = split_name(val)
        if spec:
            print(f"  Row {r}: [{val}] → 品名=[{name}] 规格=[{spec}]")
            if not dry_run:
                ws.cell(r, col_idx).value = name
                # 规格写到右边一列（如果为空）
                next_cell = ws.cell(r, col_idx + 1)
                if next_cell.value is None:
                    next_cell.value = spec

    if not dry_run:
        wb.save(path)
        print(f"\n已保存: {path}")
    else:
        print("\n[预览完成，未保存]")

    wb.close()


if __name__ == "__main__":
    main()
