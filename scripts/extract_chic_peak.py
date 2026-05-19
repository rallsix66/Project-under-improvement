
import openpyxl, os

folder = r"D:\try one\trae one\成分表处理脚本\CHIC.PEAK产品成分汇总"

# Collect all files with their category (subdirectory name)
product_files = []
for subdir in os.listdir(folder):
    subdir_path = os.path.join(folder, subdir)
    if os.path.isdir(subdir_path):
        for f in os.listdir(subdir_path):
            if (f.endswith(".xlsm") or f.endswith(".xlsx")) and not f.startswith("~$"):
                product_files.append((os.path.join(subdir_path, f), subdir, f))

product_files.sort(key=lambda x: x[2])

# ---- Color detection ----
COLORANTS = {
    "TITANIUM DIOXIDE": "白色 (White)",
    "CI 77891": "白色 (White)",
    "ZINC OXIDE": "白色 (White)",
    "KAOLIN": "白色 (White)",
    "TALC": "白色 (White)",
    "IRON OXIDE RED": "红色 (Red)",
    "IRON OXIDE YELLOW": "黄色 (Yellow)",
    "IRON OXIDE BLACK": "黑色 (Black)",
    "CI 77491": "红色 (Red)",
    "CI 77492": "黄色 (Yellow)",
    "CI 77499": "黑色 (Black)",
    "CI 77489": "橙色 (Orange)",
    "CI 77007": "蓝色 (Blue)",
    "ULTRAMARINE": "蓝色 (Blue)",
    "CI 77288": "绿色 (Green)",
    "CI 77742": "紫色 (Violet)",
    "MANGANESE VIOLET": "紫色 (Violet)",
    "CI 77019": "珠光白 (Pearly White)",
    "MICA": "珠光/透明",
    "CHARCOAL": "黑色 (Black)",
    "CI 19140": "黄色 (Yellow)",
    "CI 15985": "橙色 (Orange)",
    "CI 16035": "红色 (Red)",
    "CI 42090": "蓝色 (Blue)",
    "CI 14700": "红色 (Red)",
    "CI 15850": "红色 (Red)",
    "CI 45410": "粉色 (Pink)",
    "CI 73360": "红色 (Red)",
    "CI 77000": "银色 (Silver)",
    "SYNTHETIC FLUORPHLOGOPITE": "珠光/透明",
    "TIN OXIDE": "珠光辅助",
    "FERRIC FERROCYANIDE": "蓝色 (Blue)",
    "CI 77510": "蓝色 (Blue)",
}

FORM_KEYWORDS = {
    "LIP STICK": "唇膏/棒状",
    "LIP LINER": "唇线笔",
    "LIP GLAZE": "唇釉",
    "LIP JELLY": "唇冻/啫喱",
    "LIP MUD": "唇泥",
    "CUSHION FOUNDATION": "气垫/粉底液",
    "LIQUID BLUSH": "腮红液/液体",
    "BLUSH MUD": "腮红泥/泥状",
    "SETTING COMPACT": "蜜粉饼/粉饼",
    "PRESSED POWDER": "蜜粉饼/粉饼",
}

SCENT_KEYWORDS = [
    "FRAGRANCE", "PARFUM", "AROMA", "PERFUME", "FLAVOR",
    "LAVANDULA", "LAVENDER", "ROSA", "ROSE", "JASMINE",
    "OSMANTHUS", "VANILLA", "MINT", "MENTHA", "CITRUS",
    "PELARGONIUM", "GERANIUM", "PATCHOULI", "BERGAMOT",
    "FLOWER OIL", "FLOWER WATER", "ESSENTIAL OIL",
    "LEAF OIL", "SEED OIL", "PEEL OIL", "FRUIT OIL",
    "CHOCOLATE", "COCOA", "COFFEE", "CINNAMON", "PEACH",
    "APRICOT", "BERRY", "CHERRY", "COCONUT",
]

# Product name translations
TRANSLATIONS = {
    "CHIC.PEAK-SUPER STAY DOUBLE-END LIP STICK & LIP LINER-#CINNAMON LATTE 204": "持久双头唇膏唇线笔-#肉桂拿铁204",
    "CHIC.PEAK-SUPER STAY DOUBLE-END LIP STICK & LIP LINER-#DUSTY ROSE 201": "持久双头唇膏唇线笔-#灰粉玫瑰201",
    "CHIC.PEAK-SUPER STAY DOUBLE-END LIP STICK & LIP LINER-#MOLASSES DATE 205": "持久双头唇膏唇线笔-#糖蜜枣205",
    "CHIC.PEAK-SUPER STAY DOUBLE-END LIP STICK & LIP LINER-#NUDE APRICOT 203": "持久双头唇膏唇线笔-#裸杏203",
    "CHIC.PEAK-SUPER STAY DOUBLE-END LIP STICK & LIP LINER-#TULIP CORAL 206": "持久双头唇膏唇线笔-#郁金香珊瑚206",
    "CHIC.PEAK-SUPER STAY DOUBLE-END LIP STICK & LIP LINER-#VELVET PEACH 202": "持久双头唇膏唇线笔-#丝绒蜜桃202",
    "CHIC.PEAK SUPER STAY DOUBLE TOUCH LIP GLAZE-C01#": "持久双头唇釉-C01#",
    "CHIC.PEAK SUPER STAY DOUBLE TOUCH LIP GLAZE-C02#": "持久双头唇釉-C02#",
    "CHIC.PEAK SUPER STAY DOUBLE TOUCH LIP GLAZE-C03#": "持久双头唇釉-C03#",
    "CHIC.PEAK SUPER STAY DOUBLE TOUCH LIP GLAZE-C04#": "持久双头唇釉-C04#",
    "CHIC.PEAK SUPER STAY DOUBLE TOUCH LIP GLAZE-C05#": "持久双头唇釉-C05#",
    "CHIC.PEAK SUPER STAY DOUBLE TOUCH LIP GLAZE-C06#": "持久双头唇釉-C06#",
    "CHIC.PEAK-HYDRO-GLOSS LIP JELLY-#01": "水光亮泽唇冻-#01",
    "CHIC.PEAK-HYDRO-GLOSS LIP JELLY-#02": "水光亮泽唇冻-#02",
    "CHIC.PEAK-HYDRO-GLOSS LIP JELLY-#03": "水光亮泽唇冻-#03",
    "CHIC.PEAK-HYDRO-GLOSS LIP JELLY-#04": "水光亮泽唇冻-#04",
    "CHIC.PEAK-HYDRO-GLOSS LIP JELLY-#05": "水光亮泽唇冻-#05",
    "CHIC.PEAK-HYDRO-GLOSS LIP JELLY-#06": "水光亮泽唇冻-#06",
    "CHIC.PEAK-HYDRO-GLOSS LIP JELLY-#07": "水光亮泽唇冻-#07",
    "CHIC.PEAK-HYDRO-GLOSS LIP JELLY-#08": "水光亮泽唇冻-#08",
    "CHIC.PEAK-HYDRO-GLOSS LIP JELLY-#09": "水光亮泽唇冻-#09",
    "CHIC.PEAK SUPER STAY CLOUD VELVET LIP MUD-P01": "持久云雾丝绒唇泥-P01",
    "CHIC.PEAK SUPER STAY CLOUD VELVET LIP MUD-P02": "持久云雾丝绒唇泥-P02",
    "CHIC.PEAK SUPER STAY CLOUD VELVET LIP MUD-P03": "持久云雾丝绒唇泥-P03",
    "CHIC.PEAK SUPER STAY CLOUD VELVET LIP MUD-P04": "持久云雾丝绒唇泥-P04",
    "CHIC.PEAK SUPER STAY CLOUD VELVET LIP MUD-P05": "持久云雾丝绒唇泥-P05",
    "CHIC.PEAK SUPER STAY CLOUD VELVET LIP MUD-P06": "持久云雾丝绒唇泥-P06",
    "CHIC.PEAK SUPER STAY CLOUD VELVET LIP MUD-P07": "持久云雾丝绒唇泥-P07",
    "CHIC.PEAK SUPER STAY CLOUD VELVET LIP MUD-P08": "持久云雾丝绒唇泥-P08",
    "CHIC.PEAK SUPER STAY CLOUD VELVET LIP MUD-P09": "持久云雾丝绒唇泥-P09",
    "CHIC.PEAK MATTE & SHINE LIP GLAZEALL-DAY WEAR-S01#": "哑光&光泽唇釉全天持妆-S01#",
    "CHIC.PEAK MATTE & SHINE LIP GLAZEALL-DAY WEAR-S02#": "哑光&光泽唇釉全天持妆-S02#",
    "CHIC.PEAK MATTE & SHINE LIP GLAZEALL-DAY WEAR-S03#": "哑光&光泽唇釉全天持妆-S03#",
    "CHIC.PEAK MATTE & SHINE LIP GLAZEALL-DAY WEAR-S04#": "哑光&光泽唇釉全天持妆-S04#",
    "CHIC.PEAK MATTE & SHINE LIP GLAZEALL-DAY WEAR-S05#": "哑光&光泽唇釉全天持妆-S05#",
    "CHIC.PEAK MATTE & SHINE LIP GLAZEALL-DAY WEAR-S06#": "哑光&光泽唇釉全天持妆-S06#",
    "CHIC.PEAK SUPER STAY SKIN GLOW CUSHION FOUNDATIONS-#01 IVORY": "持久焕亮气垫粉底SPF50+ PA+++-#01象牙白",
    "CHIC.PEAK SUPER STAY SKIN GLOW CUSHION FOUNDATIONS-#02 NEUTRAL": "持久焕亮气垫粉底SPF50+ PA+++-#02自然色",
    "CHIC.PEAK SUPER STAY SKIN GLOW CUSHION FOUNDATIONS-#03 WARM SANDL": "持久焕亮气垫粉底SPF50+ PA+++-#03暖沙色",
    "CHIC.PEAK TAP-TAP PUFF LIQUID BLUSH-01": "拍拍粉扑液体腮红-01",
    "CHIC.PEAK TAP-TAP PUFF LIQUID BLUSH-02": "拍拍粉扑液体腮红-02",
    "CHIC.PEAK TAP-TAP PUFF LIQUID BLUSH-03": "拍拍粉扑液体腮红-03",
    "CHIC.PEAK TAP-TAP PUFF LIQUID BLUSH-04": "拍拍粉扑液体腮红-04",
    "CHIC.PEAK TAP-TAP PUFF LIQUID BLUSH-05": "拍拍粉扑液体腮红-05",
    "CHIC.PEAK TAP-TAP CLOUD VELVET BLUSH MUD-M01": "拍拍云雾丝绒腮红泥-M01",
    "CHIC.PEAK TAP-TAP CLOUD VELVET BLUSH MUD-M02": "拍拍云雾丝绒腮红泥-M02",
    "CHIC.PEAK TAP-TAP CLOUD VELVET BLUSH MUD-M03": "拍拍云雾丝绒腮红泥-M03",
    "CHIC.PEAK TAP-TAP CLOUD VELVET BLUSH MUD-M04": "拍拍云雾丝绒腮红泥-M04",
    "CHIC.PEAK TAP-TAP CLOUD VELVET BLUSH MUD-M05": "拍拍云雾丝绒腮红泥-M05",
    "CHIC.PEAK TAP-TAP CLOUD VELVET BLUSH MUD-M06": "拍拍云雾丝绒腮红泥-M06",
    "CHIC.PEAK TAP-TAP CLOUD VELVET BLUSH MUD-M07": "拍拍云雾丝绒腮红泥-M07",
    "CHIC.PEAK TAP-TAP CLOUD VELVET BLUSH MUD-M08": "拍拍云雾丝绒腮红泥-M08",
    "CHIC.PEAK PRISMATIC SOFT-MATTE SETTING COMPACT": "棱镜柔雾定妆蜜粉饼",
}

def get_form(name):
    forms = []
    for kw, form_type in FORM_KEYWORDS.items():
        if kw in name.upper():
            forms.append(form_type)
    if not forms:
        return "未识别"
    return " / ".join(forms)

def get_color_info(ingredients):
    colors_found = []
    for ing_name in ingredients:
        name_upper = ing_name.upper().strip()
        for color_kw, color_desc in COLORANTS.items():
            if color_kw in name_upper:
                entry = f"{color_desc} - {ing_name[:50]}"
                if entry not in colors_found:
                    colors_found.append(entry)
                break
    if not colors_found:
        return "未检测到色素成分"
    return " | ".join(colors_found)

def get_scent_info(ingredients, product_name):
    scent_ings = []
    for ing in ingredients:
        name_upper = ing["name"].upper().strip()
        for kw in SCENT_KEYWORDS:
            if kw in name_upper:
                if ing["name"] not in [s["name"] for s in scent_ings]:
                    scent_ings.append(ing)
                break

    scent_parts = []
    for si in scent_ings:
        name = si["name"]
        func = si["func"]
        if func and name.upper() != "FRAGRANCE":
            scent_parts.append(f"{name} [{func}]")
        else:
            scent_parts.append(name)

    if not scent_parts:
        return "未检测到香精/香味成分", ""

    unique = list(dict.fromkeys(scent_parts))
    scent_str = " | ".join(unique[:8])
    scent_type = determine_scent_type(product_name, scent_ings)
    return scent_type, scent_str

def determine_scent_type(name, scent_ings):
    name_upper = name.upper()

    # Color cosmetics typically unscented or lightly scented
    flavor_map = {
        "CINNAMON": "肉桂味",
        "PEACH": "蜜桃味",
        "APRICOT": "杏味",
        "ROSE": "玫瑰味",
        "CORAL": "珊瑚（无味）",
        "VANILLA": "香草味",
        "BERRY": "莓果味",
        "CHERRY": "樱桃味",
        "COFFEE": "咖啡味",
    }
    for kw, scent in flavor_map.items():
        if kw in name_upper:
            return scent

    for si in scent_ings:
        n = si["name"].upper()
        if "FRAGRANCE" in n or "PARFUM" in n or "AROMA" in n:
            return "有香精 (未标明具体香型)"

    return "无香/原料味 (Unscented)"


COLS = ["序号", "文件名", "产品名", "中文品名（翻译）", "CHIC.PEAK产品成分汇总-中文品名", "产品形状/质地", "产品颜色", "产品气味/香型", "香味成分详情", "色素成分详情", "Sheet信息"]

wb_out = openpyxl.Workbook()
ws = wb_out.active
ws.title = "产品形状颜色气味"
ws.append(COLS)

print(f"共 {len(product_files)} 个产品文件\n")

for idx, (path, category, filename) in enumerate(product_files, 1):
    try:
        wb = openpyxl.load_workbook(path, data_only=True)

        all_ingredients = []
        all_sheet_names = []
        product_name = ""

        all_sheet_names = list(wb.sheetnames)

        # Template workaround: all .xlsm files share a first sheet "LIP STICK#CINNAMON LATTE 204"
        # whose R1C1 is the template product name. Skip it; use the LAST sheet only.
        data_sheets = wb.sheetnames[1:] if len(wb.sheetnames) > 1 else wb.sheetnames

        for sn in data_sheets:
            ws_data = wb[sn]
            rows = list(ws_data.iter_rows(values_only=True))

            # Take product name from the first data sheet's R1C1
            if not product_name and rows and rows[0] and rows[0][0]:
                product_name = str(rows[0][0]).strip()

            headers = rows[1] if len(rows) > 1 else []
            inci_col = None
            func_col = None

            for i, h in enumerate(headers):
                if h is None:
                    continue
                hs = str(h).strip().lower()
                if "inci" in hs or "ingredients" in hs:
                    inci_col = i
                if "function" in hs or "目的" in hs or "作用" in hs or "使用目的" in hs:
                    func_col = i

            for row in rows[2:]:
                name = ""
                func = ""
                if inci_col is not None and len(row) > inci_col and row[inci_col]:
                    name = str(row[inci_col]).strip()
                if func_col is not None and len(row) > func_col and row[func_col]:
                    func = str(row[func_col]).strip()
                if name or func:
                    all_ingredients.append({"name": name, "func": func})

        wb.close()

        # Product name may have SPF info truncated in cell - normalize
        # Try to match translation
        cn_translated = ""
        for en_name, cn_name in TRANSLATIONS.items():
            if en_name.upper() in product_name.upper() or product_name.upper() in en_name.upper():
                cn_translated = cn_name
                break
        if not cn_translated:
            # Try exact match with file name
            file_base = os.path.splitext(filename)[0]
            for en_name, cn_name in TRANSLATIONS.items():
                if file_base.upper() in en_name.upper() or en_name.upper() in file_base.upper():
                    cn_translated = cn_name
                    break
        if not cn_translated:
            cn_translated = "需手动翻译"

        form = get_form(product_name)
        color_info = get_color_info([ing["name"] for ing in all_ingredients])
        scent_type, scent_detail = get_scent_info(all_ingredients, product_name)

        colorant_names = []
        for ing_name in [ing["name"] for ing in all_ingredients]:
            name_upper = ing_name.upper().strip()
            for color_kw in COLORANTS:
                if color_kw in name_upper:
                    if ing_name not in colorant_names:
                        colorant_names.append(ing_name[:80])
                    break
        color_detail = " | ".join(colorant_names[:10]) if colorant_names else "无"

        row_data = [idx, filename, product_name, cn_translated, category, form, color_info, scent_type, scent_detail, color_detail]
        row_data.append(" | ".join(all_sheet_names))

        ws.append(row_data)
        print(f"{idx:2d}. {str(product_name)[:55]:55s} | 品类: {category} | 气味: {scent_type}")

    except Exception as e:
        print(f"{idx:2d}. ERROR: {e}")
        continue

# Format columns
widths = [6, 55, 55, 38, 22, 28, 70, 28, 55, 60, 35]
for col_letter, w in zip("ABCDEFGHIJK", widths):
    ws.column_dimensions[col_letter].width = w

out_path = r"D:\try one\trae one\成分表处理脚本\CHIC.PEAK产品成分汇总\CHIC.PEAK产品形状颜色气味表.xlsx"
wb_out.save(out_path)
print(f"\n完成! 输出: {out_path}")
print(f"共 {ws.max_row - 1} 个产品")
