
import openpyxl, os, re

folder = r"D:\try one\trae one\成分表处理脚本\全部成分表汇总"
files = [f for f in os.listdir(folder) if f.endswith(".xlsx") and not f.startswith("~$")]
files.sort()

# ---- Build Chinese name mapping from ICE LERSKIN产品成分汇总 folder ----
cn_summary_folder = r"D:\try one\trae one\成分表处理脚本\ICE LERSKIN产品成分汇总"
cn_name_map = {}  # english_filename -> chinese product name

if os.path.exists(cn_summary_folder):
    for subdir in os.listdir(cn_summary_folder):
        subdir_path = os.path.join(cn_summary_folder, subdir)
        if os.path.isdir(subdir_path):
            for f in os.listdir(subdir_path):
                if f.endswith(".xlsm") or f.endswith(".xlsx") and not f.startswith("~$"):
                    # Match the English filename to the Chinese directory name
                    base_name = os.path.splitext(f)[0]
                    cn_name_map[base_name] = subdir

# ---- Color detection ----
COLORANTS = {
    "TITANIUM DIOXIDE": "白色 (White)",
    "CI 77891": "白色 (White)",
    "ZINC OXIDE": "白色 (White)",
    "KAOLIN": "白色 (White)",
    "IRON OXIDE RED": "红色 (Red)",
    "IRON OXIDE YELLOW": "黄色 (Yellow)",
    "IRON OXIDE BLACK": "黑色 (Black)",
    "IRON OXIDE": "氧化铁色",
    "CI 77491": "红色 (Red)",
    "CI 77492": "黄色 (Yellow)",
    "CI 77499": "黑色 (Black)",
    "CI 77489": "橙色 (Orange)",
    "CI 77007": "蓝色 (Blue)",
    "ULTRAMARINE": "蓝色 (Blue)",
    "CI 77288": "绿色 (Green)",
    "CHROMIUM OXIDE GREEN": "绿色 (Green)",
    "CI 77289": "绿色 (Green)",
    "CI 77019": "珠光白 (Pearly White)",
    "MICA": "珠光/透明",
    "CHARCOAL": "黑色 (Black)",
    "BAMBOO CHARCOAL": "黑色 (Black)",
    "CI 19140": "黄色 (Yellow)",
    "CI 15985": "橙色 (Orange)",
    "CI 16035": "红色 (Red)",
    "CI 42090": "蓝色 (Blue)",
    "CI 14700": "红色 (Red)",
    "CI 17200": "红色 (Red)",
    "CI 42053": "绿色 (Green)",
    "CI 47005": "黄色 (Yellow)",
    "CI 61570": "绿色 (Green)",
    "CI 75470": "红色 (Red)",
}

# ---- Form detection from product name ----
FORM_KEYWORDS = {
    "SERUM": "精华液/液体",
    "TONER": "化妆水/液体",
    "GEL": "凝胶",
    "CREAM": "乳霜/霜状",
    "MOUSSE": "泡沫/慕斯",
    "MASK": "面膜/泥状",
    "SPRAY": "喷雾",
    "STICK": "棒状",
    "OIL": "油状",
    "CLEANSER": "洁面/液体",
    "BATH GLOVES": "手套型/液体",
    "LIP ESSENCE": "唇部精华/液体",
    "SUNSCREEN SPRAY": "防晒喷雾",
    "SCREEN": "乳液/防晒",
    "ELIXIR": "护发油/油状",
    "TREATMENT SPRAY": "护理喷雾",
}

# ---- Scent detection ----
SCENT_INGREDIENT_KEYWORDS = [
    "FRAGRANCE", "PARFUM", "AROMA", "PERFUME", "FLAVOR",
    "LAVANDULA", "LAVENDER", "ROSA", "ROSE", "JASMINE", "JASMINUM",
    "OSMANTHUS", "FIG", "BANANA", "GRAPE", "GRAPEFRUIT", "PEACH",
    "CHAMOMILE", "ROSEMARY", "MINT", "MENTHA", "CITRUS",
    "PELARGONIUM", "GERANIUM", "YLANG", "SANDALWOOD",
    "VANILLA", "PATCHOULI", "BERGAMOT", "NEROLI",
    "FLOWER OIL", "FLOWER WATER", "ESSENTIAL OIL",
    "LEAF OIL", "SEED OIL", "PEEL OIL", "FRUIT OIL",
]

def get_form(name):
    name_upper = name.upper()
    forms = []
    for kw, form_type in FORM_KEYWORDS.items():
        if kw in name_upper:
            forms.append(form_type)
    if not forms:
        return "未识别"
    return " / ".join(forms)

def get_color_info(ingredients, sheet_names):
    """Extract color info from ingredients and sheet names."""
    colors_found = []

    for sn in sheet_names:
        sn_upper = sn.upper().strip()
        if "WHITE" in sn_upper:
            colors_found.append("白色 (White) - from sheet name")
        if "BLACK" in sn_upper:
            colors_found.append("黑色 (Black) - from sheet name")
        if "GREEN" in sn_upper:
            colors_found.append("绿色 (Green) - from sheet name")
        if "PINK" in sn_upper:
            colors_found.append("粉色 (Pink) - from sheet name")

    for ing_name in ingredients:
        name_upper = ing_name.upper().strip()
        for color_kw, color_desc in COLORANTS.items():
            if color_kw in name_upper:
                if color_desc not in colors_found:
                    colors_found.append(f"{color_desc} - {ing_name[:50]}")
                break

    if not colors_found:
        return "未检测到色素成分"
    return " | ".join(colors_found)

def get_scent_info(ingredients, product_name):
    """Extract scent info from ingredients and product name."""
    scent_ings = []

    for ing in ingredients:
        name_upper = ing["name"].upper().strip()
        for kw in SCENT_INGREDIENT_KEYWORDS:
            if kw in name_upper:
                if ing["name"] not in [s["name"] for s in scent_ings]:
                    scent_ings.append(ing)
                break

    scent_parts = []
    for si in scent_ings:
        name = si["name"]
        func = si["func"]
        if func and func != "Fragrance" and name.upper() != "FRAGRANCE":
            scent_parts.append(f"{name} [{func}]")
        else:
            scent_parts.append(name)

    if not scent_parts:
        return "未检测到香精/香味成分", ""

    unique = list(dict.fromkeys(scent_parts))
    scent_str = " | ".join(unique[:8])

    scent_type = determine_scent_type(product_name, scent_ings)

    return scent_type, scent_str

def determine_scent_type(name, scent_ings):
    """Determine the likely scent type."""
    name_upper = name.upper()

    fruit_scent_map = {
        "BANANA": "香蕉味",
        "GRAPE GLOW": "葡萄味",
        "GRAPEFRUIT": "西柚味",
        "PEACH": "水蜜桃味",
        "FIG": "无花果味",
        "JASMINE": "茉莉花香",
        "LAVENDER": "薰衣草香",
        "OSMANTHUS": "桂花香",
    }

    for kw, scent in fruit_scent_map.items():
        if kw in name_upper:
            return scent

    flower_map = {
        "ROSA DAMASCENA": "玫瑰花香 (Rose)",
        "ROSA ": "玫瑰花香 (Rose)",
        "LAVANDULA": "薰衣草香 (Lavender)",
        "JASMINUM": "茉莉花香 (Jasmine)",
        "CHAMOMILLA": "洋甘菊香 (Chamomile)",
        "OSMANTHUS": "桂花香 (Osmanthus)",
        "VIOLA": "紫罗兰香 (Violet)",
        "NEROLI": "橙花香 (Neroli)",
        "YLANG": "依兰香 (Ylang)",
        "CENTAUREA": "矢车菊香 (Cornflower)",
    }
    for si in scent_ings:
        n = si["name"].upper()
        for kw, desc in flower_map.items():
            if kw in n:
                return desc

    has_fragrance = False
    has_aroma = False
    has_parfum = False
    flower_count = 0
    herb_count = 0

    for si in scent_ings:
        n = si["name"].upper()
        if "FRAGRANCE" in n:
            has_fragrance = True
        if "AROMA" in n:
            has_aroma = True
        if "PARFUM" in n:
            has_parfum = True
        if any(kw in n for kw in ["FLOWER WATER", "FLOWER OIL", "ROSA ", "CHAMOMILE", "JASMINE", "LAVENDER", "CENTAUREA"]):
            flower_count += 1
        if any(kw in n for kw in ["ROOT EXTRACT", "LEAF EXTRACT", "ROSEMARY", "MINT", "PATCHOULI", "HERB"]):
            herb_count += 1

    if has_fragrance or has_parfum:
        if flower_count > 0:
            return "花香调 (Floral)"
        elif herb_count > 0:
            return "草本清香 (Herbal)"
        else:
            return "有香精 (未标明具体香型)"

    if has_aroma:
        if herb_count > 0:
            return "草本植物味 (Herbal Aroma)"
        if flower_count > 0:
            return "花清香 (Floral Aroma)"
        return "清香 (Aroma)"

    if flower_count > 0:
        return "花香 (Floral - natural)"

    if herb_count > 0:
        return "原料味/草本 (Natural Herbal)"

    return "无香/原料味 (Unscented)"

# ---- Product name translations ----
PRODUCT_TRANSLATIONS = {
    "Ice Lerskin 7% Glycolic Acid Toner": "7%乙醇酸爽肤水",
    "ICE LERSKIN AIR LIGHT UV SHIELDSUN STICK": "气感轻盈UV防晒棒",
    "ICE LERSKIN Acne Removing and Cleansing Mousse": "祛痘洁面慕斯",
    "ICE LERSKIN Anti-Gravity Retinol Massage Eye Cream": "抗重力视黄醇按摩眼霜",
    "Dual-Color Cleansing Clay Mask": "双色清洁泥膜",
    "ICE LERSKIN Glacier Water Essence Amino Acid Facial Cleanser": "冰川水精华氨基酸洁面乳",
    "ICE LERSKIN ICE LERSKIN SUNSPRITZ PRO AIR-LIGHT SPRAY": "气感轻盈防晒喷雾",
    "Ice Lerskin MULTI-EFFECT WHITENING SUNSCREEN SPRAY": "多效美白防晒喷雾",
    "ICE LERSKIN SEA FENNEL Youth Activating Lotion Toner": "海茴香青春活肤水",
    "ICE LERSKIN UV PLUS Aqua Shake Day Screen": "UV Plus水感摇摇防晒乳",
    "ICE LERSKIN UV PLUSAqua Shake Day Screen": "UV Plus水感摇摇防晒乳",
    "ICE LERSKIN-BHA PORE CLARIFYING GEL": "BHA毛孔净化凝胶",
    "ICE LERSKIN-C E FERULIC ANTIOXIDANT SERUM": "CE阿魏酸抗氧化精华",
    "ICE LERSKIN-C-VIT HYDRO-GLOW LIP ESSENCE-BANANA GLOW": "维C水光唇部精华-香蕉",
    "ICE LERSKIN-C-VIT HYDRO-GLOW LIP ESSENCE-GRAPE GLOW": "维C水光唇部精华-葡萄",
    "ICE LERSKIN-C-VIT HYDRO-GLOW LIP ESSENCE-GRAPEFRUIT GLOW": "维C水光唇部精华-西柚",
    "ICE LERSKIN-C-VIT HYDRO-GLOW LIP ESSENCE-PEACH GLOW": "维C水光唇部精华-蜜桃",
    "ICE LERSKIN-DAY & NIGHT BLEMISH RESCUE SERUM": "日夜祛痘急救精华",
    "ICE LERSKIN-EVEN-TONE BRIGHTENING SERUM": "匀亮焕白精华",
    "ICE LERSKIN-GLOW POLISH ESSENCE BATH GLOVES": "亮肤抛光精华沐浴手套",
    "ICE LERSKIN-TRIPLE ACID CLARIFYING TREATMENT SPRAY": "三重酸净肤护理喷雾",
    "ICE LERSKiN PRECISION ACNE SPOT SERUM": "精准祛痘点涂精华",
    "Lithe Luminous Repair Hair Elixir - Fig": "轻盈亮泽修复护发精油-无花果",
    "Lithe Luminous Repair Hair Elixir - Jasmine": "轻盈亮泽修复护发精油-茉莉",
    "Lithe Luminous Repair Hair Elixir - Lavender": "轻盈亮泽修复护发精油-薰衣草",
    "Lithe Luminous Repair Hair Elixir - Osmanthus": "轻盈亮泽修复护发精油-桂花",
}

# ---- Columns in output ----
COLS = ["序号", "文件名", "产品名", "中文品名（翻译）", "ICE LERSKIN产品成分汇总-中文品名", "产品形状/质地", "产品颜色", "产品气味/香型", "香味成分详情", "色素成分详情", "Sheet信息"]

wb_out = openpyxl.Workbook()
ws = wb_out.active
ws.title = "产品形状颜色气味"
ws.append(COLS)

for idx, f in enumerate(files, 1):
    path = os.path.join(folder, f)
    wb = openpyxl.load_workbook(path, data_only=True)

    all_ingredients = []
    all_sheet_names = []
    product_name = ""

    for sn in wb.sheetnames:
        ws_data = wb[sn]
        rows = list(ws_data.iter_rows(values_only=True))

        if not product_name and rows and rows[0] and rows[0][0]:
            product_name = str(rows[0][0]).strip()

        all_sheet_names.append(sn)

        headers = rows[1] if len(rows) > 1 else []
        inci_col = None
        func_col = None

        for i, h in enumerate(headers):
            if h is None:
                continue
            hs = str(h).strip().lower()
            if "inci" in hs or "ingredients" in hs:
                inci_col = i
            if "function" in hs or "目的" in hs or "作用" in hs or "使用目的" in hs:
                func_col = i

        if inci_col is None:
            for i, h in enumerate(headers):
                if h is None:
                    continue
                hs = str(h).strip().lower()
                if "inci" in hs:
                    inci_col = i
                if "主要使用目的" in hs or "使用目的" in hs:
                    func_col = i

        for row in rows[2:]:
            name = ""
            func = ""
            if inci_col is not None and len(row) > inci_col and row[inci_col]:
                name = str(row[inci_col]).strip()
            if func_col is not None and len(row) > func_col and row[func_col]:
                func = str(row[func_col]).strip()
            if name or func:
                all_ingredients.append({"name": name, "func": func})

    wb.close()

    # Look up Chinese product name
    file_base = os.path.splitext(f)[0]
    cn_product = cn_name_map.get(file_base, "未找到对应中文品名")

    # Look up translated product name
    cn_translated = PRODUCT_TRANSLATIONS.get(product_name, "")
    if not cn_translated:
        # Try fuzzy match
        for en_name, cn_name in PRODUCT_TRANSLATIONS.items():
            if en_name.upper() in product_name.upper() or product_name.upper() in en_name.upper():
                cn_translated = cn_name
                break
    if not cn_translated:
        cn_translated = "需手动翻译"

    # Analysis
    form = get_form(product_name)
    color_info = get_color_info([ing["name"] for ing in all_ingredients], all_sheet_names)
    scent_type, scent_detail = get_scent_info(all_ingredients, product_name)

    colorant_names = []
    for ing_name in [ing["name"] for ing in all_ingredients]:
        name_upper = ing_name.upper().strip()
        for color_kw in COLORANTS:
            if color_kw in name_upper:
                if ing_name not in colorant_names:
                    colorant_names.append(ing_name[:80])
                break

    color_detail = " | ".join(colorant_names[:8]) if colorant_names else "无"

    row_data = [idx, f, product_name, cn_translated, cn_product, form, color_info, scent_type, scent_detail, color_detail]
    row_data.append(" | ".join(all_sheet_names) if len(all_sheet_names) > 1 else all_sheet_names[0] if all_sheet_names else "")

    ws.append(row_data)
    print(f"{idx:2d}. {product_name[:45]:45s} | 中文: {cn_product}")

# Format
ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 55
ws.column_dimensions["C"].width = 50
ws.column_dimensions["D"].width = 28
ws.column_dimensions["E"].width = 28
ws.column_dimensions["F"].width = 55
ws.column_dimensions["G"].width = 30
ws.column_dimensions["H"].width = 55
ws.column_dimensions["I"].width = 50
ws.column_dimensions["J"].width = 25

out_path = r"D:\try one\trae one\成分表处理脚本\产品形状颜色气味表_v3.xlsx"
wb_out.save(out_path)
print(f"\n完成！输出: {out_path}")
print(f"ICE LERSKIN产品: {sum(1 for v in cn_name_map.values())} 个中文品名映射")
print(f"Lithe产品: {25 - sum(1 for v in cn_name_map.values())} 个（无中文映射）")
