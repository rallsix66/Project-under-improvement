# -*- coding: utf-8 -*-
"""分析基准 COA CP 产品：提取颜色/性状/气味模式，对照成分表 CI 色号"""
import os, re
import win32com.client
import openpyxl

SKILL_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BASELINE = os.path.join(SKILL_DIR, "resources", "COA基准", "多款-COA-出报告信息确认表.xls")
CP_DIR = r"E:\AI\Claude Code\成分表处理脚本\调试\CP成分"

def read_baseline_cp():
    """从基准 .xls 读取 CP sheet 所有产品"""
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = excel.Workbooks.Open(BASELINE)
    ws = wb.Worksheets("CP")

    products = []
    for r in range(7, ws.UsedRange.Rows.Count + 1):
        idx = ws.Cells(r, 1).Value
        name = ws.Cells(r, 2).Value
        spec = ws.Cells(r, 3).Value
        color_shape = ws.Cells(r, 4).Value
        odor = ws.Cells(r, 5).Value
        if idx and name:
            cn_text = str(name).split('\n')[0].strip() if name else ""
            en_text = str(name).split('\n')[1].strip() if name and '\n' in str(name) else ""
            cs_text = str(color_shape).split('\n') if color_shape else ["", ""]
            cs_cn = cs_text[0].strip()
            cs_en = cs_text[1].strip() if len(cs_text) > 1 else ""
            od_text = str(odor).split('\n') if odor else ["", ""]
            od_cn = od_text[0].strip()
            od_en = od_text[1].strip() if len(od_text) > 1 else ""
            products.append({
                'row': r, 'name_cn': cn_text, 'name_en': en_text,
                'spec': str(spec).strip() if spec else "",
                'color_shape_cn': cs_cn, 'color_shape_en': cs_en,
                'odor_cn': od_cn, 'odor_en': od_en,
            })

    wb.Close(False)
    excel.Quit()
    return products

def find_matching_file(name_en, brand='CP'):
    """在 CP成分 找匹配的 .xlsm"""
    name_key = re.sub(r'[\s\.\-\#\+]', '', name_en.upper())[:30]
    for fname in os.listdir(CP_DIR):
        if not fname.endswith('.xlsm'):
            continue
        fn_key = re.sub(r'[\s\.\-\#\+]', '', fname.upper().replace('.XLSM', ''))[:40]
        if name_key in fn_key or fn_key in name_key:
            return os.path.join(CP_DIR, fname)
        # 模糊匹配: 8+ chars common prefix
        if len(name_key) >= 8 and (name_key[:8] in fn_key or fn_key[:8] in name_key):
            return os.path.join(CP_DIR, fname)
    return None

def extract_ci_and_fragrance(xlsm_path):
    """提取 .xlsm 的 CI 码列表 和 是否有香精"""
    if not xlsm_path:
        return [], False
    try:
        wb = openpyxl.load_workbook(xlsm_path, data_only=True)
        ci_codes = []
        has_frag = False

        # Find the right sheet (skip phantom LIP STICK)
        for sn in wb.sheetnames:
            ws = wb[sn]
            a1 = str(ws.cell(1,1).value or '').upper()
            if 'CHIC' in a1 or 'LIP' in sn.upper():
                for r in range(3, ws.max_row + 1):
                    func = str(ws.cell(r, 7).value or '').lower()
                    if '香精' in func or 'fragrance' in func:
                        has_frag = True
                    for cc in [2, 3]:
                        ci = str(ws.cell(r, cc).value or '').strip()
                        m = re.match(r'(CI\s*\d+)', ci, re.I)
                        if m:
                            code = m.group(1).upper().replace(' ', '')
                            if code not in ci_codes:
                                ci_codes.append(code)
                wb.close()
                return ci_codes, has_frag
        wb.close()
    except:
        pass
    return [], False

def parse_color_shape(cs_cn):
    """拆分基准的颜色+性状，如 '浅珊瑚粉色膏状物' → ('浅珊瑚粉色', '膏状物')"""
    # 常见性状后缀
    shape_suffixes = ['膏状物', '膏状体', '膏状', '液体', '液状', '乳液', '凝胶',
                      '粉状物', '粉状', '泥状', '棒状', '粘稠液体', '烟雾剂',
                      '气雾', '喷雾', '洁面', '慕斯', '泡沫', '霜状', '乳霜',
                      '手套型', '精华']
    for suffix in shape_suffixes:
        if cs_cn.endswith(suffix):
            color_part = cs_cn[:-len(suffix)]
            if not color_part:
                color_part = cs_cn
            return color_part, suffix
    return cs_cn, ""

def main():
    products = read_baseline_cp()
    print(f"基准 CP 产品: {len(products)}")

    # 匹配成分文件并分析
    for p in products:
        xlsm = find_matching_file(p['name_en'])
        if xlsm:
            ci, has_frag = extract_ci_and_fragrance(xlsm)
        else:
            ci, has_frag = [], False

        color, shape = parse_color_shape(p['color_shape_cn'])

        p['file'] = os.path.basename(xlsm) if xlsm else 'NOT FOUND'
        p['ci_codes'] = ci
        p['has_fragrance'] = has_frag
        p['color_only'] = color
        p['shape_only'] = shape

    # 输出分析
    print(f"\n{'='*80}")
    print(f"{'品名':<25} | {'基准颜色':<16} | {'性状':<10} | {'CI码':<30} | {'文件匹配'}")
    print(f"{'='*80}")

    # 按产品类型分组统计
    color_patterns = {}
    shape_patterns = {}

    for p in products:
        name_short = p['name_cn'][:24]
        print(f"{name_short:<25} | {p['color_only']:<16} | {p['shape_only']:<10} | {','.join(p['ci_codes'])[:30]:<30} | {p['file'][:50]}")

        # 统计颜色模式
        ci_key = ','.join(sorted(p['ci_codes']))
        if ci_key not in color_patterns:
            color_patterns[ci_key] = []
        color_patterns[ci_key].append({
            'name': p['name_cn'],
            'color': p['color_only'],
            'shape': p['shape_only'],
            'odor': p['odor_cn'],
        })

    # 按 CI 码汇总颜色描述
    print(f"\n{'='*80}")
    print("按 CI 组合汇总颜色描述:")
    print(f"{'='*80}")
    for ci_key, entries in sorted(color_patterns.items(), key=lambda x: -len(x[1])):
        if not ci_key:
            ci_key = "(无CI)"
        colors = [e['color'] for e in entries]
        shapes = [e['shape'] for e in entries]
        print(f"\n  CI: {ci_key} ({len(entries)}产品)")
        for e in entries:
            print(f"    {e['name'][:40]} → {e['color']} | {e['shape']} | {e['odor']}")

if __name__ == '__main__':
    main()
