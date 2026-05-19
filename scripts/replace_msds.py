import openpyxl, os, glob, re
from copy import copy as pycopy

ref_dir = r'D:\try one\trae one\成分表处理脚本\参考表\ICE'
msds_dir = r'D:\try one\trae one\成分表处理脚本\MSDS\MSDS\ICE-MSDS'

def copy_cell_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = pycopy(src_cell.font)
        dst_cell.fill = pycopy(src_cell.fill)
        dst_cell.border = pycopy(src_cell.border)
        dst_cell.alignment = pycopy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format

def find_reference(product_name):
    for f in glob.glob(os.path.join(ref_dir, '*.xlsm')):
        wb_ref = openpyxl.load_workbook(f, keep_vba=False)
        for sn in wb_ref.sheetnames:
            ws_ref = wb_ref[sn]
            a1 = str(ws_ref.cell(row=1, column=1).value or '').strip()
            if a1 and product_name.upper() in a1.upper():
                return wb_ref, ws_ref, f, sn
        wb_ref.close()
    return None, None, None, None

# --- 只做第一个文件 ---
msds_file = os.path.join(msds_dir, 'MSDS_ICE_广州俪妍化妆品技术有限公司.xlsx')
print('Processing:', os.path.basename(msds_file))

wb_msds = openpyxl.load_workbook(msds_file)

ingredient_sheets = []
for sn in wb_msds.sheetnames:
    ws = wb_msds[sn]
    a1 = str(ws.cell(row=1, column=1).value or '')
    m = re.search(r'ICE\s+LERSKIN\s*.+', a1, re.IGNORECASE)
    if m:
        ingredient_sheets.append((sn, m.group(0).strip()))

print('Ingredient sheets:', ingredient_sheets)

for sheet_name, product_name in ingredient_sheets:
    print('\nSheet:', sheet_name, '|', product_name)

    wb_ref, ws_ref, ref_path, ref_sn = find_reference(product_name)
    if wb_ref is None:
        print('  No reference found!')
        continue

    print('  Reference:', os.path.basename(ref_path), '| sheet:', ref_sn)

    ws_msds = wb_msds[sheet_name]

    # 1. 解除MSDS原有的合并
    msds_merges = list(ws_msds.merged_cells.ranges)
    for mc in msds_merges:
        ws_msds.unmerge_cells(str(mc))
    print('  Unmerged %d MSDS ranges' % len(msds_merges))

    # 2. 读取参考表的合并单元格范围
    ref_merges = list(ws_ref.merged_cells.ranges)

    # 3. 清空需要写的区域（取最大范围，确保旧数据全清）
    need_rows = max(ws_msds.max_row, ws_ref.max_row)
    need_cols = max(ws_msds.max_column, ws_ref.max_column)
    for r in range(1, need_rows + 1):
        for c in range(1, need_cols + 1):
            ws_msds.cell(row=r, column=c).value = None

    # 4. 从参考表逐格复制值 + 样式
    for r in range(1, ws_ref.max_row + 1):
        # 复制行高
        if ws_ref.row_dimensions[r].height:
            ws_msds.row_dimensions[r].height = ws_ref.row_dimensions[r].height
        for c in range(1, ws_ref.max_column + 1):
            src = ws_ref.cell(row=r, column=c)
            dst = ws_msds.cell(row=r, column=c)
            dst.value = src.value
            copy_cell_style(src, dst)

    # 4b. 清除超出参考表范围的残留行（值+样式）
    empty_fill = openpyxl.styles.PatternFill(fill_type=None)
    empty_font = openpyxl.styles.Font()
    empty_border = openpyxl.styles.Border()
    empty_align = openpyxl.styles.Alignment()
    for r in range(ws_ref.max_row + 1, need_rows + 1):
        for c in range(1, need_cols + 1):
            cell = ws_msds.cell(row=r, column=c)
            cell.value = None
            cell.fill = pycopy(empty_fill)
            cell.font = pycopy(empty_font)
            cell.border = pycopy(empty_border)
            cell.alignment = pycopy(empty_align)

    # 5. 复制列宽
    for c in range(1, ws_ref.max_column + 1):
        letter = openpyxl.utils.get_column_letter(c)
        src_col = ws_ref.column_dimensions[letter]
        dst_col = ws_msds.column_dimensions[letter]
        if src_col.width:
            dst_col.width = src_col.width

    # 6. 添加参考表的合并单元格
    for mc in ref_merges:
        ws_msds.merge_cells(str(mc))
    print('  Added %d reference merged ranges' % len(ref_merges))

    print('  Copied %d rows x %d cols (data + style + merge)' % (ws_ref.max_row, ws_ref.max_column))
    wb_ref.close()

out = msds_file.replace('.xlsx', '_test.xlsx')
wb_msds.save(out)
print('\nSaved:', out)
