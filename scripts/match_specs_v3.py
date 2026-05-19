import openpyxl, re

# ========== 1. Read product library ==========
lib_path = r"D:\try one\trae one\成分表处理脚本\Icelerskin 商品库(供应链物流版本).xlsx"
wb_lib = openpyxl.load_workbook(lib_path, data_only=True)
ws_lib = wb_lib[wb_lib.sheetnames[0]]
lib_products = []
for row in ws_lib.iter_rows(min_row=1, values_only=True):
    name = str(row[1] or "").strip()
    if name:
        lib_products.append(name)
wb_lib.close()

def extract_spec(name):
    """Extract spec from product name."""
    patterns = [
        r'(\d+\.?\d*\s*[gG]\s*\+\s*\d+\.?\d*\s*[gG])',   # 1.1g+0.25g
        r'(\d+\.?\d*\s*[gG]\s*\*\s*\d+)',                   # 7.5g*10
        r'(\d+\.?\d*\s*[mM][lL])',                          # 150ML, 90ml
        r'(\d+\.?\d*\s*[gG])',                               # 20g, 1.8g, 100g
    ]
    for pat in patterns:
        m = re.search(pat, name)
        if m:
            return m.group(1).upper().replace(" ", "")
    return ""

# ========== 2. Read our products ==========
v7 = r"D:\try one\trae one\成分表处理脚本\产品形状颜色气味表_v7.xlsx"
wb_v7 = openpyxl.load_workbook(v7, data_only=True)

def read_our(ws):
    products = []
    headers = [str(c.value or "") for c in ws[1]]
    col_en = col_cn = col_shape = col_color = col_scent = col_family = None
    for i, h in enumerate(headers):
        if h == "产品名": col_en = i
        elif "中文品名" in h and "汇总" not in h: col_cn = i
        elif "形状" in h: col_shape = i
        elif h == "产品颜色": col_color = i
        elif h == "产品整体色系": col_family = i
        elif "气味" in h: col_scent = i
    for row in ws.iter_rows(min_row=2, values_only=True):
        en = str(row[col_en] or "").strip() if col_en is not None else ""
        cn = str(row[col_cn] or "").strip() if col_cn is not None else ""
        shape = str(row[col_shape] or "").strip() if col_shape is not None else ""
        color = str(row[col_color] or "").strip() if col_color is not None else ""
        scent = str(row[col_scent] or "").strip() if col_scent is not None else ""
        family = str(row[col_family] or "").strip() if col_family is not None else ""
        if cn or en:
            products.append((cn, en, shape, color, scent, family))
    return products

ice_products = read_our(wb_v7["产品形状颜色气味"])
cp_products = read_our(wb_v7["CHIC.PEAK产品"])
wb_v7.close()

# ========== 3. Matching logic ==========

def match_ice(en, cn):
    """Match ICE product to library. Returns (lib_name, spec)."""
    en_up = en.upper()
    best = ""
    best_spec = ""
    best_score = 0

    for lib_name in lib_products:
        lib_up = lib_name.upper()
        score = 0

        # Brand check
        is_lithe = "LITHE" in en_up and "ELIXIR" in en_up
        if is_lithe:
            if "LITHE" in lib_up:
                score += 1
        else:
            if "ICE LERSKIN" in lib_up or "ICELERSKIN" in lib_up:
                score += 1

        # === Product-specific matching ===
        if "GLYCOLIC ACID TONER" in en_up:
            if "乙醇" in lib_name and "爽肤" in lib_name: score += 10
        if "SEA FENNEL" in en_up:
            if "海茴香" in lib_name: score += 10
        if "SHIELDSUN" in en_up or "AIR LIGHT UV" in en_up:
            if "防晒棒" in lib_name: score += 10
            if "20G" in lib_up: score += 1  # Prefer 20g over 15g
        if "ACNE" in en_up and "MOUSSE" in en_up:
            if "洁面幕斯" in lib_name or "洁面慕斯" in lib_name: score += 10
        if "ANTI-GRAVITY" in en_up or ("RETINOL" in en_up and "EYE" in en_up):
            if "视黄醇" in lib_name or "眼霜" in lib_name: score += 10
            if "电动" in lib_name: score -= 1
        if "DUAL-COLOR" in en_up or "DUAL COLOR" in en_up:
            if "双色" in lib_name and "泥膜" in lib_name: score += 10
        if "GLACIER" in en_up and ("AMINO" in en_up or "CLEANSER" in en_up):
            if "氨基酸洗面奶" in lib_name or "氨基酸洁面" in lib_name: score += 10
            if "100G" in lib_up: score += 1  # Prefer version with spec
        if "SUNSPRITZ" in en_up:
            if "防晒水" in lib_name or "防晒喷雾" in lib_name: score += 8
            if "防晒水喷雾" in lib_name: score += 3
            if "ML" in lib_up: score += 1
        if "MULTI-EFFECT WHITENING" in en_up:
            if "美白" in lib_name or ("防晒喷雾" in lib_name and "升级" not in lib_name): score += 8
            if "150ML" in lib_up: score += 2
            if "升级版" in lib_name: score -= 1  # Prefer original 150ml
        if "UV PLUS" in en_up or ("AQUA" in en_up and "SHAKE" in en_up):
            if "防晒乳" in lib_name: score += 10
        if "BHA" in en_up and ("PORE" in en_up or "GEL" in en_up):
            if "BHA" in lib_up: score += 10
            if "凝胶" in lib_name: score += 3
            if "泥膜" in lib_name: score -= 10  # Penalize mud mask
        if "C E FERULIC" in en_up or ("FERULIC" in en_up and "ANTIOXIDANT" in en_up):
            if "阿魏酸" in lib_name: score += 10
            if re.search(r'\bCE\b', lib_up): score += 5
        if "HYDRO-GLOW LIP ESSENCE" in en_up:
            if "唇部精华" in lib_name or "唇精华" in lib_name: score += 10
            if "BANANA" in en_up and "香蕉" in lib_name: score += 5
            if "GRAPE" in en_up and "葡萄" in lib_name: score += 5
            if "GRAPEFRUIT" in en_up and "西柚" in lib_name: score += 5
            if "PEACH" in en_up and "蜜桃" in lib_name: score += 5
        if "DAY & NIGHT" in en_up or "BLEMISH RESCUE" in en_up:
            if "祛痘精华" in lib_name: score += 10
            if "棉棒" in lib_name: score -= 5
        if "EVEN-TONE" in en_up or "BRIGHTENING" in en_up:
            if "美白" in lib_name: score += 5
            if "精华" in lib_name: score += 3
        if "BATH GLOVES" in en_up:
            if "沐浴手套" in lib_name: score += 10
        if "TRIPLE ACID" in en_up:
            if "三重酸" in lib_name or "祛痘喷雾" in lib_name: score += 10
        if "PRECISION ACNE" in en_up:
            if "精准祛痘" in lib_name or "祛痘精华" in lib_name: score += 10
            if "棉棒" in lib_name: score -= 5
        if is_lithe:
            if "护发精油" in lib_name: score += 8
            if "套盒" in lib_name: score -= 5
            if "FIG" in en_up and "无花果" in lib_name: score += 5
            if "JASMINE" in en_up and "茉莉" in lib_name: score += 5
            if "LAVENDER" in en_up and "薰衣草" in lib_name: score += 5
            if "OSMANTHUS" in en_up and "桂花" in lib_name: score += 5

        # Prefer matches that have a spec
        spec = extract_spec(lib_name)
        if spec:
            score += 0.5

        if score > best_score:
            best_score = score
            best = lib_name
            best_spec = spec
        elif score == best_score and spec and not best_spec:
            # Same score but this one has a spec → prefer it
            best = lib_name
            best_spec = spec

    # Require minimum score to avoid false matches
    if best_score < 3:
        return ""

    # Inference: individual 护发精油 = 20ML (套盒 is 20ml*4瓶)
    if not best_spec and "ELIXIR" in en_up:
        for scent in ["FIG", "JASMINE", "LAVENDER", "OSMANTHUS"]:
            if scent in en_up:
                return "20ML"

    return best_spec


def match_cp(en, cn):
    """Match CHIC.PEAK product to library. Returns (lib_name, spec)."""
    en_up = en.upper()
    best = ""
    best_spec = ""
    best_score = 0

    for lib_name in lib_products:
        lib_up = lib_name.upper()
        score = 0

        if "CHIC" not in lib_up:
            continue

        score += 1  # Base brand score

        # S01-S06: 新版双头唇釉
        for s in ["S01", "S02", "S03", "S04", "S05", "S06"]:
            if s in en_up and s in lib_up:
                score += 8

        # C01#-C06#: 双头唇釉
        for c in ["C01", "C02", "C03", "C04", "C05", "C06"]:
            if c in en_up and c in lib_up:
                if "双头唇釉" in lib_name or "LIP GLAZE" in lib_up:
                    score += 10

        # P01-P09: 唇泥 (P01-P06 in library, P07-P09 inferred)
        for p in ["P01", "P02", "P03", "P04", "P05", "P06"]:
            if p in en_up and p in lib_up:
                if "唇泥" in lib_name:
                    score += 10

        # #01-#09: 唇冻 (#01-#06 in library, #07-#09 inferred)
        for j in ["#01", "#02", "#03", "#04", "#05", "#06"]:
            if j in en_up and j in lib_up:
                if "唇冻" in lib_name or "LIP JELLY" in lib_up:
                    score += 10

        # M01-M08: 腮红
        for m in ["M01", "M02", "M03", "M04", "M05", "M06", "M07", "M08"]:
            if m in en_up and m in lib_up:
                if "腮红" in lib_name:
                    score += 10

        # 201-206: 双头唇膏唇线笔
        for d in ["201", "202", "203", "204", "205", "206"]:
            if d in en_up and d in lib_up:
                if "唇膏" in lib_name and "唇线" in lib_name:
                    score += 10

        # 气垫 cushion
        if "CUSHION" in en_up:
            if "防晒气垫" in lib_name or "气垫" in lib_name:
                score += 8
            if "IVORY" in en_up and "象牙白" in lib_name: score += 6
            if "NEUTRAL" in en_up and ("中性" in lib_name or "自然" in lib_name): score += 6
            if "WARM SAND" in en_up and "暖" in lib_name: score += 6
            if "唇冻" in lib_name: score -= 20
            if "替换装" in lib_name: score -= 5

        # 蜜粉饼
        if "SETTING COMPACT" in en_up:
            if "粉饼" in lib_name: score += 10

        # Prefer matches with spec
        spec = extract_spec(lib_name)
        if spec:
            score += 0.5

        if score > best_score:
            best_score = score
            best = lib_name
            best_spec = spec
        elif score == best_score and spec and not best_spec:
            best = lib_name
            best_spec = spec

    # Only trust matches with sufficient score
    if best_score < 5:
        best_spec = ""

    # Inference rules for products not in library or without spec
    if not best_spec:
        # P07-P09: same series as P01-P06 → 1.8G
        for p in ["P07", "P08", "P09"]:
            if p in en_up:
                return "1.8G"
        # #07-#09: same series as #01-#06 → 2.2G
        for j in ["#07", "#08", "#09"]:
            if j in en_up:
                return "2.2G"
        # S01-S06: same series as C01-C06 → 2.5G+2.5G
        for s in ["S01", "S02", "S03", "S04", "S05", "S06"]:
            if s in en_up:
                return "2.5G+2.5G"

    return best_spec


# ========== 4. Run matching ==========
ICE_SPEC_MAP = {}
for cn, en, shape, color, scent, family in ice_products:
    spec = match_ice(en, cn)
    if spec:
        ICE_SPEC_MAP[cn] = spec

CP_SPEC_MAP = {}
for cn, en, shape, color, scent, family in cp_products:
    spec = match_cp(en, cn)
    if spec:
        CP_SPEC_MAP[cn] = spec

# ========== 5. Report ==========
lines = []
lines.append("=== ICE ===")
for cn, en, shape, color, scent, family in ice_products:
    spec = ICE_SPEC_MAP.get(cn, "")
    status = "OK" if spec else "???"
    lines.append(f"  {status} | {cn:30s} | {spec:15s}")

lines.append("\n=== CP ===")
for cn, en, shape, color, scent, family in cp_products:
    spec = CP_SPEC_MAP.get(cn, "")
    status = "OK" if spec else "???"
    lines.append(f"  {status} | {cn:30s} | {spec:15s}")

with open(r"D:\try one\trae one\spec_match_v3.txt", "w", encoding="utf-8") as f:
    f.write("\n".join(lines))

ice_found = sum(1 for v in ICE_SPEC_MAP.values() if v)
cp_found = sum(1 for v in CP_SPEC_MAP.values() if v)
print(f"ICE: {ice_found}/25, CP: {cp_found}/53")
print("Check spec_match_v3.txt")
