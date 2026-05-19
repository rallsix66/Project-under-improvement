# -*- coding: utf-8 -*-
"""COM版：生成COA（ICE+CP两个sheet），只改数据不改样式"""
import os, re, json
import openpyxl
import win32com.client

# ====== 路径配置（脚本所在 skill 目录） ======
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SKILL_DIR = os.path.dirname(SCRIPT_DIR)
TEMPLATE = os.path.join(SKILL_DIR, "templates", "COA模板.xls")
RESOURCES = os.path.join(SKILL_DIR, "resources")

# ====== 委托方信息（从 resources/新建文本文档.txt 读取） ======
def load_delegate_info():
    info_file = os.path.join(RESOURCES, "新建文本文档.txt")
    info = {}
    if os.path.exists(info_file):
        with open(info_file, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if "：" in line:
                    k, v = line.split("：", 1)
                    info[k.strip()] = v.strip()
    return info

_DELEGATE = load_delegate_info()
DELEGATE_CN = _DELEGATE.get("委托单位", "厦门冰丽研化妆品有限公司")
DELEGATE_EN = "Xiamen Bingliyan Cosmetics Co., Ltd."
DELEGATE_ADDR_CN = _DELEGATE.get("委托单位地址", "中国厦门市思明区莲前西路2号2503室A区")
DELEGATE_ADDR_EN = "Room 2503A, Block A, No. 2 Lianqian West Road, Siming District, Xiamen City, China"

# ====== 工具函数 ======
def extract_spec(text):
    for pat in [r'(\d+\.?\d*\s*[gG]\s*\+\s*\d+\.?\d*\s*[gG])',
                r'(\d+\.?\d*\s*[mM][lL]\s*\+\s*\d+\.?\d*\s*[mM][lL])',
                r'(\d+\.?\d*\s*[mM][lL]\s*\*\s*\d+\s*[片片])',
                r'(\d+\.?\d*\s*[mM][lL]\s*\*\s*\d+)',
                r'(\d+\.?\d*\s*[mM][lL])', r'(\d+\.?\d*\s*[gG])']:
        m = re.search(pat, text)
        if m: return m.group(1).replace(" ","").lower()
    return ""

# ====== CP 产品码提取 ======
def extract_cp_code(name_cn, name_en):
    """从品名提取 CP 产品编码，用于查颜色映射表"""
    combined = (name_cn + ' ' + name_en).upper()
    # 产品码格式按优先级: S01/C04/M08/P03/201/#01/01#
    for pat in [r'\b(S\d{2})\b', r'\b(C\d{2})\b', r'\b(M\d{2})\b',
                r'\b(P\d{2})\b', r'\b(\d{3})\b',
                r'(\d{2})#', r'#(\d{2})\b']:
        m = re.search(pat, combined)
        if m: return m.group(0)
    # 特殊产品
    if 'SETTING COMPACT' in combined: return 'SETTING_COMPACT'
    return None

def detect_cp_product_type(name_cn, name_en):
    """识别 CP 产品类型"""
    combined = (name_cn + ' ' + name_en).upper()
    if 'LIP GLAZEALL' in combined or ('LIP GLAZE' in combined and 'MATTE' in combined):
        return '哑光亮光双头唇釉'
    if 'LIP GLAZE' in combined and 'DOUBLE TOUCH' in combined:
        return '双头唇釉'
    if 'LIP MUD' in combined: return '唇泥'
    if 'BLUSH MUD' in combined: return '腮红泥'
    if 'LIQUID BLUSH' in combined: return '液体腮红'
    if 'LIP JELLY' in combined: return '按压唇冻'
    if 'LIP STICK' in combined and 'LIP LINER' in combined: return '双头唇膏+唇线笔'
    if 'CUSHION' in combined: return '防晒气垫'
    if 'SETTING COMPACT' in combined: return '粉饼'
    return None

# ====== CP 色号映射表 ======
def load_cp_color_map():
    map_path = os.path.join(RESOURCES, "cp_color_map.json")
    if os.path.exists(map_path):
        with open(map_path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

CP_COLOR_MAP = load_cp_color_map()
print(f"CP色号映射: {len(CP_COLOR_MAP)}条")

def read_ingredients_for_scent(ref_path, product_name_en, brand):
    """扫描成分文件的正确 sheet，返回 (has_fragrance, ci_codes)"""
    if not ref_path or not os.path.exists(ref_path):
        return False, []
    try:
        wb_ref = openpyxl.load_workbook(ref_path, data_only=True)
        name_up = product_name_en.upper().strip()
        fname_up = os.path.basename(ref_path).upper()
        ws_ref = None
        # A1 匹配
        for sn in wb_ref.sheetnames:
            ws = wb_ref[sn]
            a1 = str(ws.cell(row=1, column=1).value or '').strip().upper()
            if a1 and name_up in a1:
                ws_ref = ws
                break
        # 回退：品牌关键词
        if ws_ref is None:
            for sn in wb_ref.sheetnames:
                sn_up = sn.upper()
                # 跳过 phantom LIP STICK sheet
                if sn_up.startswith('LIP STICK'):
                    if 'LIP STICK' not in fname_up and 'LIP LINER' not in fname_up:
                        continue
                if brand == 'ICE' and ('ICE' in sn_up or 'LITHE' in sn_up or 'DUAL' in sn_up):
                    ws_ref = wb_ref[sn]; break
                elif brand == 'CP' and 'CHIC' in sn_up:
                    ws_ref = wb_ref[sn]; break
        # 回退：第一个非 phantom
        if ws_ref is None:
            for sn in wb_ref.sheetnames:
                sn_up = sn.upper()
                if sn_up.startswith('LIP STICK'):
                    if 'LIP STICK' in fname_up or 'LIP LINER' in fname_up:
                        pass
                    else:
                        continue
                ws_ref = wb_ref[sn]; break
        if ws_ref is None:
            wb_ref.close()
            return False, []

        has_frag = False; ci_codes = []
        for r in range(3, ws_ref.max_row + 1):
            func = str(ws_ref.cell(r, 7).value or '').lower()
            if '香精' in func or 'fragrance' in func:
                has_frag = True
            for cc in [2, 3]:
                ci = str(ws_ref.cell(r, cc).value or '').strip()
                m = re.match(r'(CI\s*\d+)', ci, re.I)
                if m:
                    ci_codes.append(m.group(1).upper().replace(' ', ''))
                    break
        wb_ref.close()
        return has_frag, ci_codes
    except Exception:
        return False, []

# CI色素 → 色系映射
def infer_ci_color(ci_codes):
    if not ci_codes: return ''
    joined = ' '.join(ci_codes)
    if any(c in joined for c in ['CI15850','CI45410','CI73360','CI17200']): return '红色系'
    if any(c in joined for c in ['CI15985','CI45380']): return '橙色系'
    if any(c in joined for c in ['CI19140','CI47005']): return '黄色系'
    if any(c in joined for c in ['CI42090','CI77007']): return '蓝色系'
    if any(c in joined for c in ['CI60725','CI77742']): return '紫色系'
    if any(c in joined for c in ['CI77491','CI77492','CI77499']): return '棕色系'
    if 'CI77891' in joined and len(ci_codes) == 1: return '白色'
    if len(ci_codes) >= 2: return '多色'
    return ''

# ====== 读取产品 ======
def get_products(folder, brand):
    products = []
    for fname in os.listdir(folder):
        if not fname.endswith('.xlsm'): continue
        fpath = os.path.join(folder, fname)
        fname_up = fname.upper()
        wb = openpyxl.load_workbook(fpath, data_only=True, keep_vba=False)
        sns = wb.sheetnames
        target_sn = None; target_a1 = None; ingr = 0
        # 第一轮：品牌关键词匹配
        for sn in sns:
            ws_t = wb[sn]
            a1 = str(ws_t.cell(1,1).value or '').strip()
            if not a1: continue
            a1_up = a1.upper()
            if sn.upper().startswith('LIP STICK'):
                if 'LIP STICK' not in fname_up and 'LIP LINER' not in fname_up:
                    continue
            if brand == 'ICE':
                if 'ICE' in a1_up or 'LITHE' in a1_up or 'DUAL' in a1_up:
                    target_sn = sn; target_a1 = a1; break
            else:
                if 'CHIC' in a1_up:
                    target_sn = sn; target_a1 = a1; break
        # Fallback: 第一个不含对面品牌的sheet
        if not target_sn:
            for sn in sns:
                ws = wb[sn]
                a1 = str(ws.cell(1,1).value or '').strip()
                a1_up = a1.upper()
                if not a1: continue
                if sn.upper().startswith('LIP STICK'):
                    if 'LIP STICK' in fname_up or 'LIP LINER' in fname_up:
                        pass
                    else:
                        continue
                if brand == 'ICE' and 'CHIC' in a1_up: continue
                if brand == 'CP' and 'CHIC' not in a1_up: continue
                target_sn = sn; target_a1 = a1; break
        if not target_sn:
            wb.close(); continue

        # 数成分
        ws = wb[target_sn]
        ingr = 0
        for r in range(3, ws.max_row+1):
            if ws.cell(r,3).value and str(ws.cell(r,3).value).strip(): ingr += 1

        if 'LIP STICK' in fname_up and 'LIP LINER' in fname_up:
            m = re.search(r'#?(DUSTY ROSE|VELVET PEACH|NUDE APRICOT|CINNAMON LATTE|MOLASSES DATE|TULIP CORAL)\s*(\d{3})', fname_up)
            if m:
                variant = m.group(1) + ' ' + m.group(2)
                target_a1 = re.sub(r'#[A-Z\s]+\d{3}', '#' + variant, target_a1)

        products.append({'filename':fname, 'sheet':target_sn, 'name_en':target_a1,
                         'ingredient_count':ingr, 'brand':brand, 'path':fpath})
        wb.close()
    return products

# ====== 匹配生产方 ======
def match_mfr(en_name, brand, mfr_table):
    en_up = en_name.upper()
    best = None; best_sc = 0
    for code, cn, mfr, addr in mfr_table:
        cn_up = cn.upper(); sc = 0
        if brand == 'ICE':
            if 'ICE' not in cn_up and 'LITHE' not in cn_up: continue
            if 'CHIC' in cn_up: continue
            sc = 1
            if 'GLYCOLIC' in en_up and ('乙醇酸' in cn or '果酸' in cn): sc += 10
            if '7% GLYCOLIC' in en_up and '7%' in cn: sc += 5
            if 'ETHANOLIC' in en_up and '果酸' in cn: sc += 5
            if 'AIR LIGHT' in en_up and '防晒棒' in cn: sc += 10
            if 'ACNE' in en_up and ('慕斯' in cn or '洁面' in cn): sc += 10
            if 'ANTI-GRAVITY' in en_up and '视黄醇' in cn: sc += 10
            if 'DUAL-COLOR' in en_up and '双色' in cn: sc += 10
            if 'GLACIER' in en_up and '氨基酸' in cn: sc += 10
            if 'SUNSPRITZ' in en_up and '防晒水' in cn: sc += 8
            if 'MULTI-EFFECT' in en_up and '多效' in cn: sc += 15
            elif 'MULTI-EFFECT' in en_up and ('美白' in cn or '提亮' in cn): sc += 5
            if 'SEA FENNEL' in en_up and '海茴香' in cn: sc += 10
            if 'UV PLUS' in en_up and '防晒乳' in cn: sc += 10
            if 'BHA' in en_up and ('BHA' in cn_up or '水杨酸' in cn): sc += 10
            if 'C E FERULIC' in en_up and '阿魏酸' in cn: sc += 10
            if 'HYDRO-GLOW LIP ESSENCE' in en_up and '唇' in cn:
                sc += 10
                for kw, zh in [('BANANA','香蕉'),('GRAPE','葡萄'),('GRAPEFRUIT','西柚'),('PEACH','蜜桃')]:
                    if kw in en_up and zh in cn: sc += 5
            if 'DAY & NIGHT' in en_up and ('日夜' in cn or '祛痘' in cn): sc += 10
            if 'EVEN-TONE' in en_up and ('美白' in cn or '提亮' in cn): sc += 10
            if 'BATH GLOVES' in en_up and '磨砂手套' in cn: sc += 10
            if 'TRIPLE ACID' in en_up and ('三重酸' in cn or '三酸' in cn): sc += 10
            if 'PRECISION ACNE' in en_up and '精准祛痘' in cn: sc += 10
            if 'LITHE' in en_up and '护发精油' in cn:
                sc += 8
                for kw, zh in [('FIG','无花果'),('JASMINE','茉莉'),('LAVENDER','薰衣草'),('OSMANTHUS','桂花')]:
                    if kw in en_up and zh in cn: sc += 5
            if '泰国' in cn: sc -= 3
            if '替换' in cn: sc -= 5
        else:
            if 'CHIC' not in cn_up: continue
            sc = 1
            for code_list, name_kw in [
                (['S01','S02','S03','S04','S05','S06'], '双头唇釉'),
                (['C01','C02','C03','C04','C05','C06'], '双头唇釉'),
                (['P01','P02','P03','P04','P05','P06','P07','P08','P09'], '唇泥'),
                (['#01','#02','#03','#04','#05','#06','#07','#08','#09'], '唇冻'),
                (['M01','M02','M03','M04','M05','M06','M07','M08'], '腮红'),
                (['201','202','203','204','205','206'], '唇膏'),
            ]:
                for cd in code_list:
                    if cd in en_up and cd in cn_up and name_kw in cn:
                        sc += 10
            if 'CUSHION' in en_up and '气垫' in cn:
                sc += 11
                for num in ['01','02','03']:
                    if ('#' + num) in en_up and num in cn_up: sc += 5
            if 'LIQUID BLUSH' in en_up and '液体腮红' in cn:
                sc += 10
                for num in ['01','02','03','04','05']:
                    if num in en_up and num in cn_up: sc += 5
            if 'SETTING COMPACT' in en_up and '粉饼' in cn: sc += 10
            if '替换' in cn: sc -= 5

        spe = extract_spec(en_up); spc = extract_spec(cn_up)
        if spe and spc and spe == spc: sc += 3
        elif spc: sc += 0.5

        if sc > best_sc: best_sc = sc; best = (code, cn, mfr, addr)
    return best, best_sc

def do_match(prods, mfr_table):
    matched, unmatched = [], []
    for idx, p in enumerate(prods):
        r, sc = match_mfr(p['name_en'], p['brand'], mfr_table)
        if r and sc >= 3:
            code, cn, mfr, addr = r
            spec = extract_spec(cn) or extract_spec(p['name_en'])
            matched.append({**p, 'code':code, 'name_cn':cn, 'spec':spec,
                           'manufacturer':mfr, 'address':addr, 'score':sc})
        else:
            spec = extract_spec(p['filename']) or extract_spec(p['name_en'])
            matched.append({**p, 'code':'', 'name_cn':p['name_en'], 'spec':spec,
                           'manufacturer':'待确认', 'address':'', 'score':0})
    return matched

# ====== 推断颜色/性状/气味 ======
SHAPE_MAP = {
    '液体':'Liquid','透明液体':'Transparent Liquid','无色透明液体':'Colorless Transparent Liquid',
    '泡沫':'Foam','乳霜':'Cream','膏状物':'Paste','膏状':'Paste',
    '慕斯':'Mousse','洁面':'Cleanser','凝胶':'Gel','气雾':'Aerosol','气雾剂':'Aerosol',
    '棒状':'Stick','手套型':'Bath Gloves','液状':'Liquid','乳液':'Lotion',
    '精华':'Serum','喷雾':'Spray','粘稠液体':'Viscous Liquid','泥状':'Mud',
    '粉状物':'Powder','粉状':'Powder','待确认':'TBD',
    '待确认膏状':'TBD Paste','待确认液状':'TBD Liquid',
}
COLOR_MAP = {
    '透明':'Transparent','白色':'White',
    '白色+黑色':'White & Black','黑色':'Black','半透明':'Translucent',
    '淡黄色':'Light Yellow','淡紫色':'Light Purple','淡粉色':'Light Pink',
    '淡橘色':'Light Orange','蜜桃色':'Peach','微白色':'Whitish',
    '乳白色':'Milky White','无色透明':'Colorless Transparent',
    '透明泡沫':'Transparent Foam',
    '红色系':'Red Tone','橙色系':'Orange Tone','黄色系':'Yellow Tone',
    '蓝色系':'Blue Tone','紫色系':'Purple Tone','棕色系':'Brown Tone',
    '按色号':'By Shade',
    '多色':'Multi-Color',
    '奶桃粉色':'Milky Peach Pink','草莓奶昔色':'Strawberry Milk Shake',
    '蜜瓜色':'Melon','树莓色':'Raspberry','烤奶棕色':'Toasted Brown',
    '象牙色':'Ivory','象牙白':'Ivory White','自然色':'Natural','中性色':'Neutral','暖肤色':'Warm Sand',
    '浆果色':'Berry','紫罗兰色':'Violet','冰茶色':'Iced Tea','日落色':'Sunset','朦胧粉色':'Hazy Pink',
    '尘土玫瑰色':'Dusty Rose','丝绒蜜桃色':'Velvet Peach',
    '裸杏色':'Nude Apricot','肉桂拿铁色':'Cinnamon Latte',
    '枣泥色':'Molasses Date','郁金香珊瑚色':'Tulip Coral',
    # CP 彩妆具体颜色
    '砖红棕色':'Brick-red Brown','浅粉裸色':'Light Pink Nude','裸棕色':'Nude Brown',
    '橘粉色':'Orange Pink','玫瑰粉色':'Rose Pink','莓粉色':'Berry Pink',
    '豆沙橘色':'Dusty Orange','深豆沙色':'Deep Dusty Mauve','珊瑚粉色':'Coral Pink',
    '鲜艳橘红色':'Bright Orange Red','灰玫粉色':'Grayish Pink','柔和裸粉色':'Soft Nude Pink',
    '棕橘色':'Brown Orange','豆沙色':'Dusty Mauve','砖红色':'Brick Red',
    '亮橘红色':'Bright Orange Red','正红色':'True Red','珊瑚橘色':'Coral Orange',
    '玫红色':'Rose Red','柔粉棕色':'Soft Pink Brown','浅粉色':'Light Pink',
    '浅珊瑚粉色':'Light Coral Pink','中粉色':'Medium Pink','珊瑚玫粉色':'Coral Rose Pink',
    '橘色':'Orange','鲜粉色':'Bright Pink','暖珊瑚橘色':'Warm Coral Orange',
    '暖杏色':'Warm Apricot','冷粉色':'Cool Pink','橘棕':'Orange Brown',
    '樱桃色':'Cherry','灰玫瑰裸粉色':'Dusty Rose Nude Pink','赤陶色':'Terracotta',
    '蜜桃粉色':'Peach Pink','暖沙色':'Warm Sand','透明':'Transparent',
    '待确认':'TBD',
}
ODOR_MAP = {
    '稍有气味':'Slight Odor','无':'None',
    '特征性气味':'Characteristic Odor','待确认':'TBD',
}

def get_shape_color_scent(brand, name_en, name_cn='', ref_path=''):
    """返回 (物理状态, 颜色, 气味) — Chinese。
    从成分表分析色素和香精推断，不依赖外部性状表。"""
    en = name_en.upper()

    has_frag, ci_codes = read_ingredients_for_scent(ref_path, name_en, brand)

    if brand == 'CP':
        odor = '稍有气味'

        # --- CP 查颜色映射表 ---
        code = extract_cp_code(name_cn, name_en)
        ptype = detect_cp_product_type(name_cn, name_en)

        map_entry = None
        if ptype and code:
            key = f"{ptype}:{code}"
            map_entry = CP_COLOR_MAP.get(key)

        if map_entry:
            state = map_entry['shape_cn']
            color = map_entry['color_cn']
        else:
            # --- CP 物理状态（按产品类型，与基准统一） ---
            if 'LIP GLAZE' in en:
                state = '膏状物'
            elif 'LIP MUD' in en or 'BLUSH MUD' in en:
                state = '膏状物'
            elif 'LIP JELLY' in en:
                state = '凝胶'
            elif 'LIQUID BLUSH' in en:
                state = '液体'
            elif 'LIP STICK' in en:
                state = '膏状物'
            elif 'CUSHION' in en:
                state = '乳液'
            elif 'SETTING COMPACT' in en:
                state = '粉状物'
            else:
                state = '膏状物'

            # --- CP 颜色提取 ---
            color = ''
            # 1) 从生产方中文名提取颜色（液体腮红/气垫）
            color_from_cn = re.search(r'(#[0-9]{2}|[0-9]{2}#)[（(]?\s*([一-鿿]+[色白])', name_cn)
            if color_from_cn:
                color = color_from_cn.group(2)
            # 2) LIP STICK 从英文名提取色号名，翻译
            if not color and 'LIP STICK' in en:
                lip_colors_en = {
                    'DUSTY ROSE': '尘土玫瑰色', 'VELVET PEACH': '丝绒蜜桃色',
                    'NUDE APRICOT': '裸杏色', 'CINNAMON LATTE': '肉桂拿铁色',
                    'MOLASSES DATE': '枣泥色', 'TULIP CORAL': '郁金香珊瑚色',
                }
                for ek, cn_color in lip_colors_en.items():
                    if ek in en:
                        color = cn_color; break
            # 3) 从CI色素推断色系
            if not color:
                color = infer_ci_color(ci_codes)
            if not color:
                color = '多色'

    else:
        # ICE 品牌
        state = ''
        if any(w in en for w in ['TONER','LOTION','WATER']): state = '液体'
        if 'SERUM' in en: state = '液体'
        if 'SPRAY' in en and 'STICK' not in en: state = '气雾剂' if 'SUN' in en else '液体'
        if 'STICK' in en: state = '棒状'
        if 'CREAM' in en and 'EYE' in en: state = '乳霜'
        if 'MOUSSE' in en or 'CLEANSER' in en or 'CLEANSING' in en: state = '慕斯'
        if 'GEL' in en: state = '凝胶'
        if 'MASK' in en or 'MUD' in en: state = '膏状'
        if 'BATH GLOVES' in en: state = '手套型'
        if 'OIL' in en or 'ELIXIR' in en: state = '液体'
        # ICE 唇部产品
        if 'HYDRO-GLOW LIP' in en or 'LIP' in en:
            if 'OIL' in en: state = '液体'
            elif 'BALM' in en: state = '膏状'
        if not state: state = '待确认'

        color = ''
        if any(w in en for w in ['CLEANSER','MOUSSE','CLEANSING']): color = '透明'
        elif state == '液体': color = '无色透明'
        elif state == '棒状': color = '白色'
        if not color: color = ''

        odor = '稍有气味' if has_frag else ''

        # ICE 历史数据覆盖
        if 'GLYCOLIC' in en: state, color, odor = '液体', '无色透明', '稍有气味'
        if 'AIR LIGHT' in en and 'SUN' in en: state, color, odor = '棒状', '白色', '稍有气味'
        if 'ACNE' in en and 'MOUSSE' in en: state, color, odor = '液体', '无色透明', '稍有气味'
        if 'ANTI-GRAVITY' in en: state, color, odor = '膏状物', '白色', '稍有气味'
        if 'DUAL-COLOR' in en: state, color, odor = '膏状物', '白色+黑色', '稍有气味'
        if 'GLACIER' in en: state, color, odor = '膏状物', '白色', '稍有气味'
        if 'SUNSPRITZ' in en: state, color, odor = '液体', '无色透明', '稍有气味'
        if 'MULTI-EFFECT WHITENING' in en: state, color, odor = '气雾剂', '乳白色', '稍有气味'
        if 'SEA FENNEL' in en: state, color, odor = '液体', '无色透明', '稍有气味'
        if 'UV PLUS' in en: state, color, odor = '乳液', '白色', '稍有气味'
        if 'BHA' in en and 'PORE' in en: state, color, odor = '凝胶', '半透明', '无'
        if 'C E FERULIC' in en: state, color, odor = '液体', '无色透明', '无'
        if 'HYDRO-GLOW LIP' in en:
            state = '膏状物'
            if 'BANANA' in en: color = '淡黄色'
            elif 'GRAPE' in en and 'GRAPEFRUIT' not in en: color = '淡紫色'
            elif 'GRAPEFRUIT' in en: color = '淡粉色'
            elif 'PEACH' in en: color = '蜜桃色'
            odor = '稍有气味'
        if 'DAY & NIGHT' in en: state, color, odor = '液体', '淡橘色', '稍有气味'
        if 'EVEN-TONE' in en: state, color, odor = '液体', '无色透明', '稍有气味'
        if 'BATH GLOVES' in en: state, color, odor = '手套型', '白色', '稍有气味'
        if 'TRIPLE ACID' in en: state, color, odor = '液体', '无色透明', '无'
        if 'PRECISION ACNE' in en: state, color, odor = '液体', '微白色', '稍有气味'
        if 'LITHE' in en: state, color, odor = '液体', '无色透明', '特征性气味'

        if not state: state = '待确认'
        if not odor: odor = '待确认'

    if not color: color = '待确认'
    if not state: state = '待确认'
    # 归一化气味：有香精/轻微气味 → 稍有气味
    if odor in ('轻微气味', '有香精'):
        odor = '稍有气味'
    return state, color, odor

# ====== 品名规范化 ======
def normalize_names(products, brand):
    """规范化中英文品名：品牌补空格、去规格、英文全大写+粘连拆分+品牌补全"""
    std = {'ICE': 'ICE LERSKIN', 'CP': 'CHIC.PEAK'}
    std_brand = std[brand]
    changes = []

    for p in products:
        cn = p['name_cn']
        en = p['name_en']

        # --- 中文品名 ---
        cn_new = re.sub(r'(?i)\bchic\s*\.?\s*peak\b', 'CHIC.PEAK', cn)
        cn_new = re.sub(r'(?i)\bice\s+lerskin\b', 'ICE LERSKIN', cn_new)
        cn_new = re.sub(r'(?i)\bice\s?lerskin\b', 'ICE LERSKIN', cn_new)
        cn_new = re.sub(r'(ICE LERSKIN)([一-鿿])', r'\1 \2', cn_new)
        cn_new = re.sub(r'(CHIC\.PEAK)([一-鿿])', r'\1 \2', cn_new)
        cn_new = re.sub(r'[（(]\s*\d+\.?\d*\s*(?:ml|g|kg|片)\s*(?:[+*x×]\s*\d+\.?\d*\s*(?:ml|g|kg|片))*[)）]\s*$', '', cn_new)
        cn_new = re.sub(r'\s*\d+\.?\d*\s*(?:ml|g|kg|片)\s*(?:[+*x×]\s*\d+\.?\d*\s*(?:ml|g|kg|片))*\s*$', '', cn_new)
        cn_new = re.sub(r'\s*\d+\.?\d*\s*(?:ml|g|kg|片)\s*(?:[+*x×]\s*\d+\.?\d*\s*(?:ml|g|kg|片))*\s*(?=[（(])', '', cn_new)
        cn_new = re.sub(r'\s*\d+\.?\d*\s*(?:ml|g|kg|片)\s*(?:[+*x×]\s*\d+\.?\d*\s*(?:ml|g|kg|片))*\s*-\s*', '-', cn_new)
        cn_new = re.sub(r'\s*-\s*$', '', cn_new)
        cn_new = re.sub(r'\s*[（(][^）)]*(?:特供|升级|版)[^）)]*[）)]\s*$', '', cn_new)
        cn_new = re.sub(r'\s+', ' ', cn_new).strip()

        # --- 英文品名 ---
        en_new = en
        en_new = re.sub(r'(?i)\bchic\s*\.?\s*peak\b', 'CHIC.PEAK', en_new)
        en_new = re.sub(r'(?i)\bice\s+lerskin\b', 'ICE LERSKIN', en_new)
        en_new = re.sub(r'(?i)\bice\s?lerskin\b', 'ICE LERSKIN', en_new)
        if std_brand.upper() not in en_new.upper():
            en_new = f'{std_brand} {en_new}'
        en_new = re.sub(r'(ICE LERSKIN)\s+\1', r'\1', en_new)
        en_new = re.sub(r'(CHIC\.PEAK)\s+\1', r'\1', en_new)
        en_new = re.sub(r'([a-z])([A-Z])', r'\1 \2', en_new)
        en_new = re.sub(r'([A-Z]+)([A-Z][a-z])', r'\1 \2', en_new)
        en_new = re.sub(r'([A-Z])(ALL\b)', r'\1 \2', en_new)
        en_new = en_new.upper()
        en_new = re.sub(r'#(?=\s|$)', '', en_new)
        en_new = re.sub(r'\s+', ' ', en_new).strip()

        if cn_new != cn or en_new != en:
            changes.append(f'  {brand}: CN="{cn}"→"{cn_new}"  EN="{en}"→"{en_new}"')
            p['name_cn'] = cn_new
            p['name_en'] = en_new

    return changes

# ====== COM生成COA ======
def generate_coa(ice_m, cp_m, output_path):
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False

    try:
        wb_com = excel.Workbooks.Open(TEMPLATE)

        def fill_sheet_com(ws_com, products):
            """用COM写数据，只改Value不改任何样式"""
            ws_com.Cells(2, 3).Value = f"{DELEGATE_CN}\n{DELEGATE_EN}"
            ws_com.Cells(3, 3).Value = f"{DELEGATE_ADDR_CN}\n{DELEGATE_ADDR_EN}"

            for i, p in enumerate(products):
                row = 7 + i
                name_cn = p.get('name_cn','')
                name_en = p.get('name_en','')
                state, color, odor = get_shape_color_scent(
                    p['brand'], p['name_en'], p.get('name_cn',''), p['path'])
                state_en = SHAPE_MAP.get(state, state)
                color_en = COLOR_MAP.get(color, color)
                odor_en  = ODOR_MAP.get(odor, odor)

                ws_com.Cells(row, 1).Value = i + 1
                ws_com.Cells(row, 2).Value = f"{name_cn}\n{name_en}"
                ws_com.Cells(row, 3).Value = p.get('spec','')
                ws_com.Cells(row, 4).Value = f"{color}{state}\n{color_en} {state_en}".strip()
                ws_com.Cells(row, 5).Value = f"{odor}\n{odor_en}"
                ws_com.Cells(row, 6).Value = ""
                ws_com.Cells(row, 7).Value = ""
                ws_com.Cells(row, 8).Value = "3年"
                ws_com.Cells(row, 9).Value = ""

        def extend_format(ws, data_end_row, col_count=9):
            TEMPLATE_LAST_ROW = 12
            if data_end_row <= TEMPLATE_LAST_ROW:
                return
            src = ws.Range(ws.Cells(TEMPLATE_LAST_ROW, 1), ws.Cells(TEMPLATE_LAST_ROW, col_count))
            dst = ws.Range(ws.Cells(TEMPLATE_LAST_ROW + 1, 1), ws.Cells(data_end_row, col_count))
            src.Copy()
            dst.PasteSpecial(Paste=-4122)
            excel.CutCopyMode = False

        def normalize_font(ws, data_end_row, col_count=9):
            from collections import Counter
            sizes = Counter()
            sample_rng = ws.Range(ws.Cells(2, 1), ws.Cells(data_end_row, col_count))
            for cell in sample_rng:
                if cell.Value is not None:
                    sizes[cell.Font.Size] += 1
            data_sizes = {s: c for s, c in sizes.items() if s < 20}
            best_size = max(data_sizes, key=data_sizes.get) if data_sizes else 11
            rng = ws.Range(ws.Cells(2, 1), ws.Cells(data_end_row, col_count))
            rng.Font.Name = "宋体"
            rng.Font.Size = best_size
            # 以 row8 字体颜色为准，统一 row7 及以下全部数据行
            ref_color = ws.Cells(8, 1).Font.Color
            data_rng = ws.Range(ws.Cells(7, 1), ws.Cells(data_end_row, col_count))
            data_rng.Font.Color = ref_color
            # 统一数据行行高（参照 row8 高度）
            ref_height = ws.Rows(8).RowHeight
            for r in range(7, data_end_row + 1):
                ws.Rows(r).RowHeight = ref_height
            title_rng = ws.Range(ws.Cells(1, 1), ws.Cells(1, col_count))
            title_rng.Font.Name = "宋体"

        # 1. 先复制模板sheet → CP
        ws1 = wb_com.Worksheets(1)
        ws1.Copy(None, ws1)
        ws_cp = wb_com.Worksheets(wb_com.Worksheets.Count)
        ws_cp.Name = "CP"

        # 2. 原sheet → ICE
        ws_ice = wb_com.Worksheets(1)
        ws_ice.Name = "ICE"

        # 3. 填ICE数据
        print("\n=== ICE ===")
        fill_sheet_com(ws_ice, ice_m)
        extend_format(ws_ice, 7 + len(ice_m) - 1)
        normalize_font(ws_ice, 7 + len(ice_m) - 1)
        for i, p in enumerate(ice_m):
            s,c,o = get_shape_color_scent(p['brand'], p['name_en'], p.get('name_cn',''), p['path'])
            print(f"  {i+1}. {p['name_cn'][:40]} | {p['spec']} | {c}{s} | {o}")

        # 4. 填CP数据
        print("\n=== CP ===")
        fill_sheet_com(ws_cp, cp_m)
        extend_format(ws_cp, 7 + len(cp_m) - 1)
        normalize_font(ws_cp, 7 + len(cp_m) - 1)
        for i, p in enumerate(cp_m):
            s,c,o = get_shape_color_scent(p['brand'], p['name_en'], p.get('name_cn',''), p['path'])
            print(f"  {i+1}. {p['name_cn'][:40]} | {p['spec']} | {c}{s} | {o}")

        wb_com.SaveAs(output_path)
        print(f"\n完成! 保存到: {output_path}")
        print(f"ICE: {len(ice_m)}产品 | CP: {len(cp_m)}产品")

    finally:
        wb_com.Close(False)
        excel.Quit()

# ====== Main ======
if __name__ == '__main__':
    import sys
    if len(sys.argv) < 3:
        print("用法: python coa_fill.py <数据目录> <输出路径>")
        print("  数据目录: 含 CP成分/、ICE成分/、厦门外贸商品对应生产方.xlsx")
        print("  输出路径: COA .xls 输出文件路径")
        sys.exit(1)

    BASE = sys.argv[1]
    OUTPUT = sys.argv[2]

    # 读生产方表
    mfr_path = os.path.join(BASE, "厦门外贸商品对应生产方.xlsx")
    if not os.path.exists(mfr_path):
        print(f"错误: 找不到生产方对照表 {mfr_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(mfr_path, data_only=True)
    ws = wb.active
    MFR = []
    for r in range(2, ws.max_row+1):
        code = str(ws.cell(r,1).value or '').strip()
        name = str(ws.cell(r,2).value or '').strip()
        mfr  = str(ws.cell(r,3).value or '').strip()
        addr = str(ws.cell(r,4).value or '').strip()
        if name: MFR.append((code,name,mfr,addr))
    wb.close()
    print(f"生产方表: {len(MFR)}条")

    # 读取产品
    ice_prods = get_products(os.path.join(BASE, "ICE成分"), 'ICE')
    cp_prods  = get_products(os.path.join(BASE, "CP成分"),  'CP')
    print(f"ICE: {len(ice_prods)}个, CP: {len(cp_prods)}个")

    # 匹配生产方
    ice_m = do_match(ice_prods, MFR)
    cp_m  = do_match(cp_prods, MFR)
    print(f"ICE匹配: {len(ice_m)}, CP匹配: {len(cp_m)}")

    # 品名规范化
    norm_changes = normalize_names(ice_m, 'ICE') + normalize_names(cp_m, 'CP')
    if norm_changes:
        print(f"\n品名修正 {len(norm_changes)} 处：")
        for ch in norm_changes:
            print(ch)

    # 检查模板是否存在
    if not os.path.exists(TEMPLATE):
        print(f"错误: 找不到COA模板 {TEMPLATE}")
        sys.exit(1)

    # 生成COA
    generate_coa(ice_m, cp_m, OUTPUT)
