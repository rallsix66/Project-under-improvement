"""
步骤⑥：CAS号填写。

填充规则：
  1. 只填写空白（None 或空字符串）的 CAS 单元格
  2. 已有值（含报送码格式 XXXXXX-XXXXX-XXXX）→ 不动
  3. "/" 或 "-" → 不动（不适用）
  4. CAS 列合并且空白 → 拆分合并，每个成分行各自填 CAS
  5. 水/AQUA/WATER → 跳过
  6. 香精/FRAGRANCE/AROMA/PARFUM → 填 "Mixture"
  7. 无 CAS 的聚合物/混合物 → 跳过（填 "/" 或不填）

用法：
  python fill_cas.py <file_or_dir> [--dry-run]
"""

import openpyxl
import os
import sys
import re
import json
import glob


# ============================================================
# CAS 映射表（INCI名称 → CAS编号）
# ============================================================
CAS_MAP = {
    # === 基础成分 ===
    'WATER': '7732-18-5',
    'AQUA': '7732-18-5',
    'SILICA': '7631-86-9',
    'BUTYLENE GLYCOL': '107-88-0',
    'PHENOXYETHANOL': '122-99-6',
    'GLYCERIN': '56-81-5',
    'HYDROXYACETOPHENONE': '99-93-4',
    '1,2-HEXANEDIOL': '6920-22-5',
    'PROPYLENE GLYCOL': '57-55-6',
    'ALCOHOL': '64-17-5',
    'STEARIC ACID': '57-11-4',
    'PALMITIC ACID': '57-10-3',
    'BHT': '128-37-0',
    'TOCOPHEROL': '10191-41-0',
    'TOCOPHERYL ACETATE': '7695-91-2',
    'BISABOLOL': '515-69-5',
    'SODIUM CHLORIDE': '7647-14-5',
    'SODIUM HYDROXIDE': '1310-73-2',
    'PARAFFIN': '8002-74-2',
    'SYNTHETIC WAX': '8002-74-2',
    'ZINC OXIDE': '1314-13-2',
    'TITANIUM DIOXIDE': '13463-67-7',
    'ALUMINUM HYDROXIDE': '21645-51-2',
    'TRIETHANOLAMINE': '102-71-6',
    'SALICYLIC ACID': '69-72-7',
    'ETHYLHEXYLGLYCERIN': '70445-33-9',
    'CAPRYLIC/CAPRIC TRIGLYCERIDE': '73398-61-5',
    'PROPANEDIOL': '504-63-2',
    'PENTYLENE GLYCOL': '5343-92-0',
    'GLYCERYL STEARATE': '31566-31-1',
    'CETEARYL ALCOHOL': '67762-27-0',
    'DIMETHICONE': '9006-65-9',
    'KAOLIN': '1332-58-7',
    'TREHALOSE': '99-20-7',
    'ALLANTOIN': '97-59-6',
    'NIACINAMIDE': '98-92-0',
    'PANTHENOL': '81-13-0',
    'SODIUM HYALURONATE': '9067-32-7',
    'BENZYL ALCOHOL': '100-51-6',
    'CHLORPHENESIN': '104-29-0',
    'FRUCTOSE': '7660-25-5',
    'PEARL POWDER': '91080-33-8',
    'MINERAL OIL': '8042-47-5',
    'ISOHEXADECANE': '93685-80-4',
    'CERA MICROCRISTALLINA': '63231-60-7',
    'MICA': '12001-26-2',
    'TALC': '14807-96-6',
    'CI 77891': '13463-67-7',
    'CI 77492': '20344-49-4',
    'CI 77491': '1309-37-1',
    'CI 77499': '12227-89-3',
    'CI 77007': '12769-96-9',
    'CI 77288': '1308-38-9',
    'CI 15850': '5281-04-9',
    'CI 45410': '18472-87-2',
    'CI 42090': '3844-45-9',
    'CI 47005': '8004-92-0',
    'CI 61570': '8021-99-6',
    'CI 19140': '1934-21-0',
    'SORBITAN OLIVATE': '223706-40-9',

    # === 防腐剂 ===
    'METHYLPARABEN': '99-76-3',
    'PROPYLPARABEN': '94-13-3',
    'ETHYLHEXYLGLYCERIN': '70445-33-9',
    'SODIUM BENZOATE': '532-32-1',
    'POTASSIUM SORBATE': '24634-61-5',

    # === 防晒剂 ===
    'ETHYLHEXYL METHOXYCINNAMATE': '5466-77-3',
    'ETHYLHEXYL SALICYLATE': '118-60-5',
    'PHENYLBENZIMIDAZOLE SULFONIC ACID': '27503-81-7',
    '4-METHYLBENZYLIDENE CAMPHOR': '36861-47-9',
    'DIETHYLAMINO HYDROXYBENZOYL HEXYL BENZOATE': '302776-68-7',
    'BIS-ETHYLHEXYLOXYPHENOL METHOXYPHENYL TRIAZINE': '187393-00-6',

    # === 硅油/乳化剂 ===
    'CYCLOPENTASILOXANE': '541-02-6',
    'METHICONE': '9004-73-3',
    'DIMETHICONE CROSSPOLYMER': '213629-14-2',
    'DIMETHICONOL': '31692-79-2',
    'TRIMETHYLSILOXYSILICATE': '56275-01-5',
    'POLYMETHYLSILSESQUIOXANE': '68554-70-1',
    'VINYL DIMETHICONE/METHICONE SILSESQUIOXANE CROSSPOLYMER': '153668-87-2',
    'ISODODECANE': '31807-55-3',
    'BIS-HYDROXYETHOXYPROPYL DIMETHICONE': '628723-36-4',
    'DIPHENYLSILOXY PHENYL TRIMETHICONE': '352230-22-9',
    'PEG/PPG-18/18 DIMETHICONE': '68937-55-3',
    'C12-20 ACID PEG-8 ESTER': '68908-68-9',
    'CETYL PEG/PPG-10/1 DIMETHICONE': '144243-53-8',
    'PEG-10 DIMETHICONE': '68937-54-2',
    'PEG-100 STEARATE': '9004-99-3',
    'TRIETHOXYCAPRYLYLSILANE': '2943-75-1',
    'DISTEARDIMONIUM HECTORITE': '97280-96-1',
    'PROPYLENE CARBONATE': '108-32-7',
    'DICAPRYLYL CARBONATE': '1680-31-5',
    'LAURETH-7': '3055-97-8',
    'C13-14 ISOPARAFFIN': '246538-79-4',
    'ETHYLHEXYL PALMITATE': '29806-73-3',
    'COCO-GLUCOSIDE': '141464-42-8',
    'POLYACRYLAMIDE': '9003-05-8',
    'MAGNESIUM ALUMINUM SILICATE': '12174-11-7',
    'XANTHAN GUM': '11138-66-2',
    'CARBOMER': '9003-01-4',
    'AMMONIUM POLYACRYLOYLDIMETHYL TAURATE': '123774-88-7',
    'HYDROXYETHYL ACRYLATE/SODIUM ACRYLOYLDIMETHYL TAURATE COPOLYMER': '111286-86-3',
    'SODIUM POLYACRYLOYLDIMETHYL TAURATE': '166432-54-6',

    # === 溶剂/推进剂 ===
    'BUTANE': '106-97-8',
    'PROPANE': '74-98-6',
    'ISOBUTANE': '75-28-5',

    # === 植物提取物 ===
    'ALOE BARBADENSIS LEAF EXTRACT': '85507-69-3',
    'PHRAGMITES COMMUNIS EXTRACT': '84604-02-4',
    'MALVA SYLVESTRIS (MALLOW) FLOWER/LEAF/STEM EXTRACT': '84082-57-5',
    'GENTIANA SCABRA ROOT EXTRACT': '95193-52-5',
    'AVENA SATIVA (OAT) KERNEL OIL': '84012-26-0',
    'POLYGONUM CUSPIDATUM ROOT EXTRACT': '501-36-0',
    'GLYCINE MAX (SOYBEAN) OIL': '8001-22-7',
    'SCUTELLARIA BAICALENSIS ROOT EXTRACT': '94279-99-9',
    'CAMELLIA SINENSIS LEAF EXTRACT': '84650-60-2',
    'GLYCYRRHIZA GLABRA (LICORICE) ROOT EXTRACT': '84775-66-6',
    'CHAMOMILLA RECUTITA (MATRICARIA) FLOWER EXTRACT': '84082-60-8',
    'ROSMARINUS OFFICINALIS (ROSEMARY) LEAF EXTRACT': '84604-14-8',
    'ZEA MAYS CORN SILK EXTRACT': '977071-11-6',
    'ZEA MAYS SILK EXTRACT': '977071-11-6',
    'ZEA MAYS (CORN) SILK EXTRACT': '977071-11-6',
    'CRAMBE ABYSSINICA SEED OIL': '68956-68-3',
    'CAMELLIA JAPONICA SEED OIL': '223748-13-8',
    'CHRYSANTHELLUM INDICUM EXTRACT': '223748-24-1',
    'ROSA DAMASCENA FLOWER WATER': '90106-38-0',
    'ASTRAGALUS MEMBRANACEUS ROOT EXTRACT': '94166-93-5',
    'CALENDULA OFFICINALIS FLOWER EXTRACT': '84776-23-8',
    'ALBIZIA JULIBRISSIN FLOWER EXTRACT': '440103-71-9',
    'GASTRODIA ELATA ROOT EXTRACT': '62499-27-8',
    'OLEA EUROPAEA (OLIVE) LEAF EXTRACT': '84012-27-1',
    'HIBISCUS ESCULENTUS FRUIT EXTRACT': '91723-07-8',
    'JASMINUM SAMBAC (JASMINE) LEAF CELL EXTRACT': '91770-17-8',
    'CENTELLA ASIATICA EXTRACT': '84696-21-9',
    'SOLUBLE COLLAGEN': '9007-34-5',
    'HYDROLYZED COLLAGEN': '92113-31-2',
    'CERAMIDE NP': '100403-19-8',

    # === 肽类 ===
    'NONAPEPTIDE-1': '158563-45-2',
    'DIPEPTIDE DIAMINOBUTYROYL BENZYLAMIDE DIACETATE': '823202-99-9',
    'ACETYL HEXAPEPTIDE-8': '616204-22-9',

    # === 其他 ===
    'LACTOBIONIC ACID': '96-82-2',
    'PENTAERYTHRITYL TETRA-DI-T-BUTYL HYDROXYHYDROCINNAMATE': '6683-19-8',
    'HYDROXYPROPYL CYCLODEXTRIN': '128446-35-5',
    'HYDROXYPHENYL PROPAMIDOBENZOIC ACID': '697235-49-7',
    'QUATERNIUM-73': '15763-48-1',
    'DIETHOXYETHYL SUCCINATE': '26962-29-8',
    'ETHYLHEXYL PALMITATE': '29806-73-3',

    # === 无 CAS（填 Mixture 或留空） ===
    'FRAGRANCE': 'Mixture',
    'AROMA': 'Mixture',
    'PARFUM': 'Mixture',
    # 无 CAS 的聚合物/混合物 → 空字符串 = 跳过
    'ACRYLATES/DIMETHICONE COPOLYMER': '',
    'ACRYLATES/AMMONIUM METHACRYLATE COPOLYMER': '',
    'ACRYLATES/C10-30 ALKYL ACRYLATE CROSSPOLYMER': '',
    'LAURYL PEG-9 POLYDIMETHYLSILOXYETHYL DIMETHICONE': '',
    'MACADAMIA SEED OIL POLYGLYCERYL-6 ESTERS BEHENATE': '',
    'EUGLENA GRACILIS POLYSACCHARIDE': '',
    'MAGNOLIA SIEBOLDII EXTRACT': '',
    'SELAGINELLA TAMARISCINA EXTRACT': '',
    'SAPOSHNIKOVIA DIVARICATA ROOT EXTRACT': '',
    'PAEONIA LACTIFLORA ROOT EXTRACT': '',
    'REHMANNIA GLUTINOSA ROOT EXTRACT': '',
    'METHYLPROPANEDIOL': '',
}


def normalize(s):
    """移除换行、空格后大写，用于模糊匹配"""
    if not s:
        return ''
    return s.replace('\n', '').replace(' ', '').upper().strip()


def is_blank(v):
    """判断 CAS 值是否为空白（需要填充）"""
    if v is None:
        return True
    s = str(v).strip()
    return s == ''


def is_protected(v):
    """判断 CAS 值是否已有内容不需填充（报送码/CAS号/不适用标记）"""
    if v is None:
        return False
    s = str(v).strip()
    if s == '':
        return False
    # "/" 或 "-" = 不适用，保护不动
    if s in ('/', '-'):
        return True
    # 报送码格式：XXXXXX-XXXXX-XXXX（数字-数字-数字）
    if re.match(r'^\d{6}-\d{5}-\d{4}$', s):
        return True
    # 已有 CAS 或其他值
    return True


def find_cas_col(ws):
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(row=2, column=c).value or '').upper()
        if 'CAS' in h:
            return c
    return None


def find_inci_col(ws):
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(row=2, column=c).value or '').upper()
        if 'INCI' in h or 'INGREDIENT' in h or '成分名称' in h:
            return c
    # Fallback: 找"标准 standard"列（中文格式的INCI列）
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(row=2, column=c).value or '')
        if '标准' in h or 'STANDARD' in h.upper():
            return c + 1  # INCI 通常在"标准"列的右边一列
    return None


def fill_cas_for_sheet(ws):
    cas_col = find_cas_col(ws)
    inci_col = find_inci_col(ws)
    if not cas_col:
        return 0, 0, 0, 0, 'no CAS column'
    if not inci_col:
        return 0, 0, 0, 0, 'no INCI column'

    # H列合并区域中，首行已有CAS → 不拆不填（厂商内部编码），首行空白 → 拆开各填各的
    h_merge_skip = set()
    for mc in list(ws.merged_cells.ranges):
        if mc.min_row >= 3 and mc.min_col <= cas_col <= mc.max_col and mc.min_row != mc.max_row:
            first_cas = ws.cell(mc.min_row, cas_col).value
            if not is_blank(first_cas):
                for rr in range(mc.min_row, mc.max_row + 1):
                    h_merge_skip.add(rr)
            else:
                ws.unmerge_cells(str(mc))

    # 辅助：安全写入 cell（处理 MergeCell 问题）
    def safe_write(r, c, val):
        cell = ws.cell(r, c)
        try:
            cell.value = val
            return True
        except AttributeError:
            # MergedCell 不可写，需要找到合并区域并解除
            for mc in list(ws.merged_cells.ranges):
                if mc.min_row <= r <= mc.max_row and mc.min_col <= c <= mc.max_col:
                    ws.unmerge_cells(str(mc))
                    break
            try:
                ws.cell(r, c).value = val
                return True
            except AttributeError:
                return False

    filled = 0
    skipped = 0
    not_found = 0
    mixture = 0
    fail = 0

    for r in range(3, ws.max_row + 1):
        cas_val = ws.cell(r, cas_col).value
        inci_val = str(ws.cell(r, inci_col).value or '').strip()

        # 跳过已被保护的行
        if is_protected(cas_val):
            skipped += 1
            continue

        # 跳过公式
        if inci_val.startswith('='):
            continue

        # 跳过空行
        if not inci_val:
            continue

        # 跳过 H 列合并且首行已有 CAS 的后续空白行
        if r in h_merge_skip:
            skipped += 1
            continue

        # 跳过水
        inci_norm = normalize(inci_val)
        if inci_norm in ('AQUA', 'WATER'):
            continue

        # 匹配 CAS
        matched = None
        # 1. 精确规范化匹配
        for key, cas in CAS_MAP.items():
            if normalize(key) == inci_norm:
                matched = cas
                break
        # 2. 子串匹配（如 TITANIUM DIOXIDE 匹配 TITANIUMDIOXIDE）
        if matched is None:
            for key, cas in CAS_MAP.items():
                key_norm = normalize(key)
                if key_norm and (key_norm in inci_norm or inci_norm in key_norm):
                    matched = cas
                    break

        if matched is not None:
            if matched == 'Mixture':
                if safe_write(r, cas_col, 'Mixture'):
                    mixture += 1
                else:
                    fail += 1
            elif matched == '':
                skipped += 1  # 无CAS的聚合物，跳过
            else:
                if safe_write(r, cas_col, matched):
                    filled += 1
                else:
                    fail += 1
        else:
            not_found += 1

    return filled, skipped, not_found, mixture, None, fail


def process_file(path, dry_run=False):
    wb = openpyxl.load_workbook(path)
    total = {'filled': 0, 'skipped': 0, 'not_found': 0, 'mixture': 0, 'fail': 0}
    details = []

    for sname in wb.sheetnames[1:]:  # 跳过 Sheet1
        ws = wb[sname]
        f, s, n, m, err, fail = fill_cas_for_sheet(ws)
        if err:
            details.append(f'  {sname}: {err}')
        elif f > 0 or n > 0 or m > 0 or fail > 0:
            details.append(f'  {sname}: filled={f}, skipped={s}, mixture={m}, not_found={n}' + (f' fail={fail}' if fail else ''))
        total['filled'] += f
        total['skipped'] += s
        total['not_found'] += n
        total['mixture'] += m
        total['fail'] += fail

    if not dry_run:
        wb.save(path)
    wb.close()
    return total, details


def main():
    if len(sys.argv) < 2:
        print('用法: python fill_cas.py <file_or_dir> [--dry-run]')
        sys.exit(1)

    path = sys.argv[1]
    dry_run = '--dry-run' in sys.argv

    if dry_run:
        print('[预览模式] 不保存\n')

    # 收集文件
    if os.path.isfile(path):
        files = [path]
    else:
        files = sorted(glob.glob(os.path.join(path, '**', '*.xlsx'), recursive=True))
        files = [f for f in files if not os.path.basename(f).startswith('~')]

    if not files:
        print(f'未找到 .xlsx 文件: {path}')
        sys.exit(1)

    grand_total = {'filled': 0, 'skipped': 0, 'not_found': 0, 'mixture': 0, 'fail': 0}

    for i, f in enumerate(files):
        fname = os.path.basename(f)
        print(f'[{i+1}/{len(files)}] {fname}')
        total, details = process_file(f, dry_run=dry_run)
        for d in details:
            print(d)
        if total['filled'] + total['not_found'] + total['mixture'] > 0:
            print(f'  >>> filled={total["filled"]}, mixture={total["mixture"]}, skipped={total["skipped"]}, not_found={total["not_found"]}')
        for k in grand_total:
            grand_total[k] += total[k]

    print(f'\n=== 总计 ===')
    print(f'  filled={grand_total["filled"]}, mixture={grand_total["mixture"]}, skipped={grand_total["skipped"]}, not_found={grand_total["not_found"]}' + (f', fail={grand_total["fail"]}' if grand_total['fail'] else ''))

    if dry_run:
        print('\n[预览完成，未保存]')


if __name__ == '__main__':
    main()
