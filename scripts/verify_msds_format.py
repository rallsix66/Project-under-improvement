# -*- coding: utf-8 -*-
"""MSDS 格式校对脚本：对比新生成文件与基准文件的格式差异"""
import os, sys
import openpyxl

def compare_format(ref_path, new_path):
    """对比两个 MSDS 文件的格式，返回差异列表"""
    issues = []
    ref_name = os.path.basename(ref_path)
    new_name = os.path.basename(new_path)

    wb_ref = openpyxl.load_workbook(ref_path)
    wb_new = openpyxl.load_workbook(new_path)

    # 1. Sheet 数量与名称
    ref_sheets = wb_ref.sheetnames
    new_sheets = wb_new.sheetnames
    if ref_sheets != new_sheets:
        issues.append(f"Sheet 列表不同: 基准={ref_sheets}, 当前={new_sheets}")
        wb_ref.close(); wb_new.close()
        return issues

    for sn in ref_sheets:
        ws_ref = wb_ref[sn]
        ws_new = wb_new[sn]
        prefix = f"{sn}: "

        # 2. 合并单元格
        ref_merges = set(str(m) for m in ws_ref.merged_cells.ranges)
        new_merges = set(str(m) for m in ws_new.merged_cells.ranges)
        if ref_merges != new_merges:
            only_ref = ref_merges - new_merges
            only_new = new_merges - ref_merges
            if only_ref: issues.append(prefix + f"缺少合并单元格: {sorted(only_ref)}")
            if only_new: issues.append(prefix + f"多余合并单元格: {sorted(only_new)}")

        # 3. 逐单元格对齐方式
        align_issues = []
        for r in range(1, ws_ref.max_row + 1):
            for c in range(1, ws_ref.max_column + 1):
                ra = ws_ref.cell(r, c).alignment
                na = ws_new.cell(r, c).alignment
                if str(ra) != str(na):
                    align_issues.append(f"({r},{c}) 基准={ra} 当前={na}")
        if align_issues:
            issues.append(prefix + f"对齐差异 {len(align_issues)} 处: {align_issues[:5]}...")

        # 4. 边框检查（只查右边界）
        border_issues = []
        for r in range(1, ws_ref.max_row + 1):
            for c in range(1, ws_ref.max_column + 1):
                rb = ws_ref.cell(r, c).border.right
                nb = ws_new.cell(r, c).border.right
                if str(rb) != str(nb):
                    border_issues.append(f"({r},{c}) 基准右框={rb} 当前={nb}")
        if border_issues:
            issues.append(prefix + f"边框差异 {len(border_issues)} 处: {border_issues[:5]}...")

        # 5. 超链接检查（仅 Sheet1 Row 13）
        if sn == ref_sheets[0]:
            ref_links = {}
            new_links = {}
            for c in range(3, ws_ref.max_column + 1):
                rc = ws_ref.cell(13, c)
                nc = ws_new.cell(13, c)
                if rc.hyperlink:
                    ref_links[c] = str(rc.hyperlink.location)
                if nc.hyperlink:
                    new_links[c] = str(nc.hyperlink.location)
            if ref_links != new_links:
                issues.append(prefix + f"超链接不一致: 基准={len(ref_links)}个, 当前={len(new_links)}个")

        # 6. 列宽
        ref_widths = {c: ws_ref.column_dimensions[c].width for c in ws_ref.column_dimensions}
        new_widths = {c: ws_new.column_dimensions[c].width for c in ws_new.column_dimensions}
        if ref_widths != new_widths:
            diff_cols = [c for c in set(list(ref_widths.keys()) + list(new_widths.keys()))
                         if ref_widths.get(c) != new_widths.get(c)]
            if diff_cols:
                issues.append(prefix + f"列宽差异: {diff_cols}")

    wb_ref.close()
    wb_new.close()
    return issues


if __name__ == '__main__':
    BASE = r"D:\try one\trae one\成分表处理脚本\调试"
    ref_dir = f"{BASE}/MSDS基准"
    out_dir = f"{BASE}/MSDS输出"

    total = 0
    for ref_name in sorted(os.listdir(ref_dir)):
        if not ref_name.endswith('.xlsx'):
            continue
        ref_path = os.path.join(ref_dir, ref_name)
        # 找对应输出文件（可能在 ICE/ 或 CP/ 子目录）
        new_path = None
        for sub in ['ICE', 'CP']:
            p = os.path.join(out_dir, sub, ref_name)
            if os.path.exists(p):
                new_path = p
                break
        if not new_path:
            print(f"  SKIP: {ref_name} (输出文件不存在)")
            continue

        issues = compare_format(ref_path, new_path)
        total += 1
        if issues:
            print(f"  DIFF: {ref_name}")
            for i in issues:
                print(f"    {i}")
        else:
            print(f"  OK: {ref_name}")

    print(f"\n检查 {total} 个文件")
